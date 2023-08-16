from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from default.models import Usuarios, Personas, Consultores, Experiencias, ExperienciasConsultor, Instituciones, Estudios, ManeraPago, TipoMoneda, Idiomas, IdiomasConsultor, Modulos, Submodulos, NivelesConocimiento, ConocimientosConsultor, CursosConsultor, NotificationConsultor, NotificationAdministrador, Empresas, EntrevistasConsultoresProyecto, Proyectos, ProyectoConsultor, NotificationEmpresa,RequerimientosModulosProyecto, EmpresaProyecto, RequerimientosIdiomasProyecto, PostulacionesProyectoGnosis, ProyectoConsultor, CategoriasConsultor, Contratos, TipoContrato, DatosBancarios, Bancos, NivelesConocimiento, NotasGnosisConsultor, Facturas, ReporteHoras, ReporteFinalActividades, TipoContrataciones
from django.contrib.auth.decorators import login_required
import math
from django.db.models import Q
from django.contrib import messages
from decimal import Decimal
import requests
from datetime import datetime
import calendar
from urllib.parse import urlencode
from .config import cipher_suite
from cryptography.fernet import Fernet
from django.contrib.auth import login, logout, authenticate
import os, re, json, ast
from datetime import date
from django.http import HttpResponseRedirect
from django.core.paginator import Paginator
from pytz import timezone
from django.conf import settings
from django.http import FileResponse
from django.urls import reverse
from django.http import HttpResponseNotFound
from num2words import num2words
# CREATE PDFS
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.views.generic import ListView
from django.shortcuts import get_object_or_404
from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
import shutil
from django.core.mail import EmailMessage
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.core.mail import send_mail
from openpyxl import Workbook
from forex_python.converter import CurrencyRates
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from calendar import monthrange


@login_required
def miConsultorProfile(request, id=None):
    try:
        user = Usuarios.objects.get(correo=request.session.get('username'))
        informationPersonalAdministradorUser = Personas.objects.get(pk=user.id_persona_id)
        
        # consultor = Usuarios.objects.get(pk=68)
        consultor = Usuarios.objects.get(id_persona=int(id))
        
        if consultor.is_superuser == 1:
            redirect('principalAdmin')
        
        informationPersonalUser = Personas.objects.get(pk=consultor.id_persona_id)
        informationConsultorUser = Consultores.objects.get(id_persona=informationPersonalUser)
        
        proyectos = ProyectoConsultor.objects.filter(id_consultor=informationConsultorUser.id)
        notas = NotasGnosisConsultor.objects.filter(id_consultor=informationConsultorUser.id)




        if proyectos.exists():  # Verificar si hay proyectos
            cantidad_proyectos = proyectos.count()
            
            puntuacion = 0

            for proyecto in proyectos:
                puntos = proyecto.puntuacion
                if puntos is not None:  # Verificar si el valor no es None
                    puntuacion += float(puntos)

            if cantidad_proyectos != 0:  # Asegurarse de que no se esté dividiendo por cero
                puntuacionConsultor = puntuacion / cantidad_proyectos
            else:
                puntuacionConsultor = 0
                cantidad_proyectos = 0
        else:
            puntuacionConsultor = 0
            cantidad_proyectos = 0

        notasValues = []
        if notas.exists():
            for nota in notas:
                notas = nota.nota   
                fecha = obtener_mes(nota.fecha_creacion)
                notasValues.append([notas, fecha])
        else:
            notasValues = 0

        # print(informationPersonalUser)
        edad = calcular_edad(str(informationPersonalUser.fecha_nacimiento))
        maneraPago2 = ManeraPago.objects.get(
            pk=informationConsultorUser.id_manera_pago_id)

        tipoMoneda = TipoMoneda.objects.get(
            pk=informationConsultorUser.id_tipo_moneda_id)
        codigo_moneda = tipoMoneda.tipo.split('(')[1].split(')')[0]

        experienciasConsultor2 = ExperienciasConsultor.objects.filter(id_consultor=informationConsultorUser)
        if experienciasConsultor2:
            myExperience = []
            a=0
            for experience in experienciasConsultor2:
                id = experience.id
                descripcion = experience.descripcion
                empresa = experience.empresa
                puesto = experience.puesto
                experienciaInicio = obtener_mes(experience.fecha_entrada)
                experienciaTermino = obtener_mes(experience.fecha_salida)
                if experience.tiempo_experiencia == 'Sigue Trabajando' or experience.tiempo_experiencia == '0':
                    experienciaTermino = 'En proceso'
                a=a+1
                myExperience.append([id, descripcion, empresa, puesto, experienciaInicio, experienciaTermino, a])
        else: 
            myExperience = []

        estudiosConsultor = Estudios.objects.filter(
            id_consultor=informationConsultorUser).order_by('-id').first()

        if estudiosConsultor:
            institucion = Instituciones.objects.get(pk=estudiosConsultor.id_institucion_id)
            educacionInicio = obtener_mes(estudiosConsultor.fecha_ingreso)

            if estudiosConsultor.fecha_termino == 'Sigue Estudiando':
                educacionTermino = 'En proceso'
            elif len(estudiosConsultor.fecha_termino) == 10 and estudiosConsultor.fecha_termino[4] == '-' and estudiosConsultor.fecha_termino[7] == '-':
                # 2021-12-17
                componentes = estudiosConsultor.fecha_termino.split('-')
                meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

                educacionTermino = meses[int(
                    componentes[1])-1] + ' ' + componentes[0]
            else:
                fecha = datetime.strptime(
                    estudiosConsultor.fecha_termino, '%Y-%m-%d %H:%M:%S.%f')
                educacionTermino = obtener_mes(fecha)
        else:
            institucion = ''
            educacionInicio = ''
            educacionTermino = ''

        ruta = 'media/PDF/' + informationPersonalUser.pais + '/' + \
            informationConsultorUser.rfc + '/' + informationConsultorUser.rfc
        cv = ruta+'_CV.pdf'
        ine = ruta+'_INE.pdf'
        actaNac = ruta+'_Acta_Nacimiento.pdf'
        pasaporte = ruta+'_Pasaporte.pdf'
        domicilio = ruta+'_Comprobante_Domicilio.pdf'
        recomendacion = ruta+'_Carta_Recomendacion.pdf'
        f3 = ruta+'_F3.pdf'
        monedaCobro = list(TipoMoneda.objects.values())
        maneraPago = list(ManeraPago.objects.values())
        listaIdiomas = list(Idiomas.objects.values())
        listaModulos = list(Modulos.objects.values())
        listaSubodulos = list(Submodulos.objects.values())
        listaNivelesConocimiento = list(NivelesConocimiento.objects.values())
        listaCategorias = list(CategoriasConsultor.objects.values())
        # misIdiomas = list(IdiomasConsultor.objects.filter(id_consultor_id=informationConsultorUser.id))

        misIdiomas = IdiomasConsultor.objects.filter(
            id_consultor_id=informationConsultorUser.id)
        myLenguajes = []
    
        misModulos = ConocimientosConsultor.objects.filter(id_consultor_id=informationConsultorUser.id)
        myModuls = []


        myCourses = []
        c=0  
        cursosConsultor = CursosConsultor.objects.filter(id_consultor=informationConsultorUser)
        if cursosConsultor:
            for cursos in cursosConsultor:
                id = cursos.id
                nombre_curso = cursos.nombre_curso
                enlace_certificado = cursos.enlace_certificado
                descripcion = cursos.descripcion
                fecha_termino = obtener_mes(cursos.fecha_termino)
                institucionCurso = Instituciones.objects.get(pk=cursos.id_institucion_curso_id)
                institucionNombre = institucionCurso.nombre
                c=c+1
                myCourses.append([id, nombre_curso, enlace_certificado, descripcion, fecha_termino, institucionNombre,  c]) 
        else:
            myCourses = []

        if misIdiomas:
            for idioma in misIdiomas:
                id = idioma.id
                nombre_idioma = Idiomas.objects.get(pk=idioma.id_idioma_id)
                nivel = idioma.nivel.split()[1]
                myLenguajes.append([id, nombre_idioma, nivel])
        else:
            myLenguajes = []
        
        c = 0
        if misModulos:
            for moduloConsulta in misModulos:
                id = moduloConsulta.id
                modulo = Modulos.objects.get(pk=moduloConsulta.id_modulo_id)
                submodulo = Submodulos.objects.get(pk=moduloConsulta.id_submodulo_id)
                nivel = NivelesConocimiento.objects.get(pk=moduloConsulta.id_nivel_id)
                c = c+1
                nivelGnosis = NivelesConocimiento.objects.get(pk=moduloConsulta.id_nivelGnosis_id)
                valido = moduloConsulta.estatus
                myModuls.append([id, modulo, submodulo, nivel, c, nivelGnosis, valido])
        else:
            myModuls = []
        
        for p in myModuls:
            # print(p[3].nombre)
            # print(p[5].nombre)
            pass
        notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')
        # //Pagination

        # //Cantidad de notificaciones que apareceran antes de crear otra page
        if notification_list.exists():
            paginator = Paginator(notification_list, 6)
            page = request.GET.get('page')
            notification = paginator.get_page(page)

            # // CANT. NOTIFICATIONS PENDING
            pending_total= NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')

        # // ONLY 4 PENDING 
            pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')[:4]
        else:
            notification = []
            pending = []
            pending_total = []

        return render(request, 'profileConsultor.html', {
            'titulo': 'Mi perfil',
            'indice': 'Consultores',
            'image':consultor.image,
            'informationPersonalUser': informationPersonalUser,
            'correo': consultor.correo,
            'edad': edad,
            'maneraPago2': maneraPago2,
            'informationConsultorUser': informationConsultorUser,
            'maneraPago': maneraPago,
            'tipoMoneda': codigo_moneda,
            'myExperience':myExperience,
            'estudiosConsultor': estudiosConsultor,
            'cursosConsultor': cursosConsultor,
            'institucion': institucion,
            'myCourses':myCourses,
            'educacionInicio': educacionInicio,
            'educacionTermino': educacionTermino,
            'monedaCobro': monedaCobro,
            'maneraPago': maneraPago,
            'listaIdiomas': listaIdiomas,
            'misIdiomas': myLenguajes,
            'listaModulos':listaModulos,
            'listaSubodulos':listaSubodulos,
            'proyectos':proyectos,
            'cantidad_proyectos':cantidad_proyectos,
            'puntuacionConsultor':puntuacionConsultor,
            'listaNivelesConocimiento':listaNivelesConocimiento,
            'misModulos':myModuls,
            'listaCategorias':listaCategorias,
            'notas':notasValues,
            'notifications':notification,
            'pending':pending,
            'pending_total':pending_total,
            'files': [
                searchFile(cv), searchFile(ine), searchFile(actaNac), searchFile(
                    pasaporte), searchFile(domicilio), searchFile(recomendacion), searchFile(f3),
            ]
        })
    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Personas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')


@login_required
def principal(request):
    try:
        user = Usuarios.objects.get(correo=request.session.get('username'))
        informationPersonalUser = Personas.objects.get(pk=user.id_persona_id)
        # print(informationPersonalUser.id)
        notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')
        # //Pagination
        
        # //Cantidad de notificaciones que apareceran antes de crear otra page
        if notification_list.exists():
            paginator = Paginator(notification_list, 6)
            page = request.GET.get('page')
            notification = paginator.get_page(page)

            # // CANT. NOTIFICATIONS PENDING
            pending_total= NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')

        # // ONLY 4 PENDING 
            pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')[:4]
        else:
            notification = []
            pending = []
            pending_total = []
        
        listaIdiomas = list(Idiomas.objects.values())
        listaModulos = list(Modulos.objects.values())
        listaSubodulos = list(Submodulos.objects.values())
        listaNivelesConocimiento = list(NivelesConocimiento.objects.values())

        return render(request, 'principalAdministradores.html',{
            'informationPersonalUser': informationPersonalUser,
            'indice':'Principal',
            'notifications':notification,
            'pending':pending,
            'pending_total':pending_total,
            'listaIdiomas': listaIdiomas,
            'listaModulos':listaModulos,
            'listaSubodulos':listaSubodulos,
            'listaNivelesConocimiento':listaNivelesConocimiento,
        })

    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Personas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')


@login_required
def updateModulosSAPAdmin(request):
    if request.method == 'POST':
        try:
            consultorID = request.POST.get('saber', '')
            nivelConsultor = request.POST.get('nivelConsultor', '')
            nivelGnosis = request.POST.get('nivelGnosis', '')
            con = request.POST.get('con', '')
            # print(consultorID)
            
            informationConsultorUser = Consultores.objects.get(pk=int(con))
            # print(informationConsultorUser)
            nivelesConocimientoConsultor = NivelesConocimiento.objects.get(pk=int(nivelConsultor))
            nivelesConocimientoGnosis = NivelesConocimiento.objects.get(pk=int(nivelGnosis))
            # print(nivelesConocimientoConsultor.nombre)
            # print(nivelesConocimientoGnosis.nombre)
            if consultorID and nivelConsultor and nivelGnosis:
                
                consulta = ConocimientosConsultor.objects.filter(
                    Q(pk=consultorID) & Q(id_consultor=informationConsultorUser))
                
                if consulta.exists():
                        
                    objeto_consulta = consulta.first()
                    # print(objeto_consulta.id)
                    objeto_consulta.id_nivel = nivelesConocimientoConsultor
                    objeto_consulta.id_nivelGnosis = nivelesConocimientoGnosis
                    objeto_consulta.estatus = 'Validado'
                    objeto_consulta.save()
                    
                    newNotificacion = NotificationConsultor(name='Gnosis SC', email='gnosis@gmail.com', subject='Validacion exitosa de tu experiencia', message='Hemos validado con exito tus modulos de experiencia, hecha un vistaso en tu perfil', status='Pending', id_consultor_destinatary_id=informationConsultorUser.id)
                    newNotificacion.save()

            return HttpResponse(status=200)
        except Consultores.DoesNotExist:
            return HttpResponse(status=500)
        except Usuarios.DoesNotExist:
            return HttpResponse(status=500)
        except Personas.DoesNotExist:
            return HttpResponse(status=500)
        except Exception as e:
            print(e)
            return HttpResponse(status=500)




@login_required
def all_notifications(request):
    try:
        user = Usuarios.objects.get(correo=request.session.get('username'))
        informationPersonalUser = Personas.objects.get(pk=user.id_persona_id)

        notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary_id=informationPersonalUser.id)
        if notification_list.exists():
        
            notification_list = notification_list.order_by('-created_at')
            paginator = Paginator(notification_list, 25)
            #print(paginator)
            page = request.GET.get('page')
            notification = paginator.get_page(page)

            # CANT. NOTIFICATIONS PENDING
            pending_total = NotificationAdministrador.objects.filter(status='Pending', id_persona_destinatary_id=informationPersonalUser.id).order_by('-created_at')

            # ONLY 4 PENDING
            pending = NotificationAdministrador.objects.filter(status='Pending', id_persona_destinatary_id=informationPersonalUser.id).order_by('-created_at')[:4]
            
            return render(request, 'notifications/all_notificationsAdministrador.html', {
                'informationPersonalUser': informationPersonalUser,
                'indice': 'Principal',
                'notification':notification,
                'notification_pending': pending,
                'pending':pending,
                'pending_total':pending_total
            })
        else:
            notification_list = []
            return render(request, 'notifications/all_notificationsAdministrador.html', {
                'informationPersonalUser': informationPersonalUser,
                'indice': 'Principal',
                'notifications': [],
                'notification_pending': [],
                'pending': [],
                'pending_total': [],
            })


    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Personas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')




# //  VIEW 1 NOTIFICATION
@login_required
def view_notification(request, id=None):
    try:
        user = Usuarios.objects.get(correo=request.session.get('username'))
        informationPersonalUser = Personas.objects.get(pk=user.id_persona_id)
       
        # // DATA DE LA NOTIFICACION ELEGIDA
        notification = NotificationAdministrador.objects.get(pk=id, id_persona_destinatary=informationPersonalUser.id)

        if notification.data:
            data = notification.data
            pares = data.split(',')
            datos = {}
            for par in pares:
                clave, valor = par.split(':')
                datos[clave] = valor

            consultor =datos['id_consultor']
            proyecto =datos['id_proyecto']
        else:
            consultor =''
            proyecto =''
        # // EL STATUS CAMBIA A 'READ' Y GUARDA
        notification.status = 'Read'
        notification.save()
        # // NOTIFICATIONS DATA
        notification_data = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')
        # // CANT. NOTIFICATIONS PENDING
        pending_total= NotificationAdministrador.objects.filter(status='Pending', id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')

        # // ONLY 4 PENDING 
        pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')[:4]
        context = {
            'indice': 'Principal',
            'informationPersonalUser': informationPersonalUser,
            'consultor':consultor,
            'proyecto':proyecto,
            'notification':notification,
            'notifications':notification_data,
            'pending':pending,
            'pending_total':pending_total,
        }
        return render(request, "notifications/view_notificationAdministrador.html", context)


    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Personas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')



# -----

@login_required
def validarEntrevista(request, id:int, proyecto:int):
    try:
        # print("Holaaaaaa")
        try:
            entrevistasConsultoresProyecto = EntrevistasConsultoresProyecto.objects.get(id_proyecto_id=int(proyecto), id_consultor_id=int(id))
            if entrevistasConsultoresProyecto:
        
                empresaUser = entrevistasConsultoresProyecto.id_empresa.id_usuario

                empresa = Empresas.objects.get(id_usuario=empresaUser)
                monedaCobro = list(TipoMoneda.objects.values())
                consultor = Consultores.objects.get(pk=entrevistasConsultoresProyecto.id_consultor_id)
                persona = Personas.objects.get(pk=consultor.id_persona_id)
                proyecto = Proyectos.objects.get(pk=entrevistasConsultoresProyecto.id_proyecto_id)
                edad = calcular_edad(str(persona.fecha_nacimiento))


                user = Usuarios.objects.get(correo=request.session.get('username'))
                informationPersonalUser = Personas.objects.get(pk=user.id_persona_id)
                informationPersonalAdministradorUser = Personas.objects.get(pk=user.id_persona_id)


                notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')
                # //Pagination
                # //Cantidad de notificaciones que apareceran antes de crear otra page
                if notification_list.exists():
                    paginator = Paginator(notification_list, 6)
                    page = request.GET.get('page')
                    notification = paginator.get_page(page)

                    # // CANT. NOTIFICATIONS PENDING
                    pending_total= NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')

                # // ONLY 4 PENDING 
                    pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')[:4]
                else:
                    notification = []
                    pending = []
                    pending_total = []


                return render(request, 'validate_interview.html',{
                    'indice': 'Proyectos',
                    'empresa':empresa,
                    'monedaCobro': monedaCobro,
                    'consultor':consultor,
                    'persona':persona,
                    'proyecto':proyecto,
                    'edad':edad,
                    'notifications':notification,
                    'pending':pending,
                    'pending_total':pending_total,
                })      
            else:
                messages.error(
                    request, 'No se encontró la entrevista para validar.')
                return redirect('principalAdmin')
        except EntrevistasConsultoresProyecto.DoesNotExist as error:
            print(f"Error: {str(error)}")
            # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
            messages.error(
                request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
            return redirect('principalAdmin')

    except Consultores.DoesNotExist as error:
            print(f"Error: {str(error)}")
            # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
            messages.error(
                request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
            return redirect('principalAdmin')
    except Empresas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('principalAdmin')
    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('principalAdmin')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('principalAdmin')


@login_required
def addConsultorProyecto(request):
    try:
        consultor = request.POST['consultor']
        consultor = Consultores.objects.get(pk=int(consultor))
        proyectoId = request.POST['proyecto']
        proyecto = Proyectos.objects.get(pk=int(proyectoId))
        moneda = request.POST['moneda']
        moneda = TipoMoneda.objects.get(pk=int(moneda))
        pago = request.POST['pago']
        fecha_inicio = request.POST['date-start']
        fecha_termino = request.POST['date-end']
        dias_seleccionados = []
        if 'lun' in request.POST:
            dias_seleccionados.append('Lun')
        if 'mar' in request.POST:
            dias_seleccionados.append('Mar')
        if 'mier' in request.POST:
            dias_seleccionados.append('Mié')
        if 'jue' in request.POST:
            dias_seleccionados.append('Jue')
        if 'vie' in request.POST:
            dias_seleccionados.append('Vie')
        if 'sab' in request.POST:
            dias_seleccionados.append('Sáb')
        if 'dom' in request.POST:
            dias_seleccionados.append('Dom')

        dias = ", ".join(dias_seleccionados)
        # print("Hola")
        horas = request.POST['inicio-horas']
        minutos = request.POST['inicio-minutos']
        cadena = horas +':'+minutos
        hora_str, minutos_str = cadena.split(':')
        horario_hora_inicio = datetime.strptime(hora_str + ':' + minutos_str, '%H:%M').time() 
        horas_finales = request.POST['final-horas']
        minutos_finales = request.POST['final-minutos']
        cadena_final = horas_finales +':'+minutos_finales
        hora_str_final, minutos_str_final = cadena_final.split(':')
        horario_hora_final = datetime.strptime(hora_str_final + ':' + minutos_str_final, '%H:%M').time() 
        fun_laborales = request.POST['funciones']

        query = ProyectoConsultor.objects.filter(id_consultor=consultor)
        
        if not query:
            
            proyectoConsultor = ProyectoConsultor(fecha_inicio=fecha_inicio, fecha_termino=fecha_termino, horario_hora_inicio=horario_hora_inicio, horario_hora_final=horario_hora_final, fun_laborales=fun_laborales, dias_laborales=dias, tarifa=pago, id_tipo_moneda=moneda, id_proyecto=proyecto, id_consultor=consultor)
            proyectoConsultor.save()

            persona = Personas.objects.get(pk=consultor.id_persona.id)
            persona.disponible = 0
            persona.save()

            entrevista = EntrevistasConsultoresProyecto.objects.get(id_proyecto=proyecto, id_consultor=consultor)
            # print(entrevista)
            entrevista.delete()

            messageEmpresa = 'Hemos validado al consultor ' + str(consultor.id_persona.nombre) + ', para trabajar en tu proyecto ' + str(proyecto.proyecto_nombre)
            empresaNotificacion = NotificationEmpresa(name='Gnosis SC', email='gnosis@gnosis.com.mx', subject='Entrevista validada', message=messageEmpresa, status='Pending', id_empresa_destinatary=entrevista.id_empresa)
            empresaNotificacion.save()
            data = 'id_proyectoConsultor:'+str(proyectoConsultor.id)
            messageConsultor = 'La empresa ' + str(entrevista.id_empresa.empresa) + ' ha aprobado tu perfil para el proyecto ' + str(proyecto.proyecto_nombre)
            newNotificacion = NotificationConsultor(name='Gnosis SC', email='gnosis@gmail.com', subject='¡¡Felicidades estas en un nuevo proyecto!!', message=messageConsultor, status='Pending', id_consultor_destinatary_id=consultor.id,  ruta='confirmarProyectoConsultor', confirm=True, data=data)
            newNotificacion.save()
            # print("envieee")
        
        url = reverse('principalAdmin')
        return redirect(url)

    except Exception as error:
        print(f"Error: {str(error)}")
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')



@login_required
def getConsultoresDisponiblesWithFilters(request):
    if request.method == 'POST':
        try:
            moduloFilter = request.POST.get('modulo-filter', '')
            submoduloFilter = request.POST.get('submodulo-filter', '')
            nivelFilter = request.POST.get('nivel-filter', '')
            tarifaFilter = request.POST.get('tarifa-filter', '')
            disponibleFilter = request.POST.get('disponible-filter', '')
            NdisponibleFilter = request.POST.get('Ndisponible-filter', '')
            nivelInglesFilter = request.POST.get('nivelIngles-filter', '')
            nivelEstudiosFilter = request.POST.get('nivelEstudios-filter', '')
            certificadosFilter = request.POST.get('certificados-filter', '')
            # print("---- Datos recibidos ----")
            # print("mod " + moduloFilter)
            # print("sub " + submoduloFilter)
            # print("nivel " + nivelFilter)
            # print("tarifa " + tarifaFilter)
            # print("disp " + disponibleFilter)
            # print("ndisp " + NdisponibleFilter)
            # print("ingle " + nivelInglesFilter)
            # print("estudi " + nivelEstudiosFilter)
            # print("certif " + certificadosFilter)

            # print("\n")

            if disponibleFilter:
                # print("disponible")
                personas = Personas.objects.filter(disponible=int(disponibleFilter))

            if NdisponibleFilter:
                # print("no dispo")
                personas = Personas.objects.filter(disponible=int(NdisponibleFilter))
            # print("\n")

            consultores = Consultores.objects.filter(id_persona__in=personas, tarifa_hora__lte=int(tarifaFilter))
            
            if not moduloFilter and not submoduloFilter and not nivelFilter:
                # print("no hay modulo aplicados")
                consultoresModulosSAP = []
            else:
                # print("hay modulos aplicados")
                conocimientosConsultor = ConocimientosConsultor.objects.filter(id_consultor__in=consultores)
                if moduloFilter:
                    # print("modulo")
                    modulo = Modulos.objects.get(pk=int(moduloFilter))
                    conocimientosConsultor = conocimientosConsultor.filter(id_modulo=modulo)
                # print("\n")
                if submoduloFilter:
                    # print("submo")
                    submodulo = Submodulos.objects.get(pk=int(submoduloFilter))
                    conocimientosConsultor = conocimientosConsultor.filter(id_submodulo=submodulo)
                # print("\n")
                if nivelFilter:
                    # print("nivelk")
                    nivelK = NivelesConocimiento.objects.get(pk=int(nivelFilter))
                    conocimientosConsultor = conocimientosConsultor.filter(id_nivel=nivelK)
                # print("\n")
                

                consultoresModulosSAP = []
                for modulosSAP in conocimientosConsultor:
                    consultoresR = Consultores.objects.filter(pk=modulosSAP.id_consultor_id)
                    # # print(consultores)
            
                consultores = Consultores.objects.filter(id__in=conocimientosConsultor.values_list('id_consultor_id', flat=True))
            

            resultadosIngles = []
            if nivelInglesFilter:
                # print("nivel imgles")
                for consultor in consultores:
                    try:
                        nivelIngles = IdiomasConsultor.objects.get(id_consultor=consultor.id, id_idioma=3, nivel=nivelInglesFilter)
                        # # print("si")
                        resultadosIngles.append(nivelIngles)
                        # Resto de tu lógica aquí
                    except IdiomasConsultor.DoesNotExist:
                        # print("no")
                        pass
                # print(resultadosIngles)
                consultores = Consultores.objects.filter(id__in=[resultado.id_consultor.id for resultado in resultadosIngles])
            else:
                # print("no hay nivel ingles")
                pass

            # print("\n")

            resultadosEstudios = []
            if nivelEstudiosFilter:
                # print("estudios educativos")
                for consultor in consultores:
                    try:
                        estudios = Estudios.objects.get(id_consultor=consultor.id, educacion=nivelEstudiosFilter)
                        # print("si")
                        resultadosEstudios.append(estudios)
                        # Resto de tu lógica aquí
                    except Estudios.DoesNotExist:
                        pass
                # print(resultadosIngles)
                consultores = Consultores.objects.filter(id__in=[resultado.id_consultor.id for resultado in resultadosEstudios])
            else:
                # print("no hay nivel estudios")
                pass
            # print("\n")


            resultadoscertificicaciones = []
            if certificadosFilter:
                # print("certificados")
                for consultor in consultores:
                    try:
                        certificados = CursosConsultor.objects.get(id_consultor=consultor.id, nombre_curso__icontains=certificadosFilter)
                        # print("si")
                        resultadoscertificicaciones.append(certificados)
                        # Resto de tu lógica aquí
                    except CursosConsultor.DoesNotExist:
                        # print("no")
                        pass
                # print(resultadosIngles)
                consultores = Consultores.objects.filter(id__in=[resultado.id_consultor.id for resultado in resultadoscertificicaciones])
            else:
                # print("no hay nivel certificados")
                pass
            # print("\n")
            # print(consultores)

        
            
            correos = []
            for consultor in consultores:
                # print(consultor.id_persona)
                usuarios = Usuarios.objects.filter(id_persona=consultor.id_persona, is_superuser=0)
                if usuarios.exists():
                    correo = usuarios[0].correo
                    correos.append(correo)

            usuariosPersona = Usuarios.objects.filter(correo__in=correos)
            #print(usuariosPersona)
            
            personasResultado = []
            for usuario in usuariosPersona:
                persona = Personas.objects.get(pk=usuario.id_persona_id)
                personasResultado.append(persona)

            resultados = []
            # print(resultados)
            for persona in personasResultado:
                img = Usuarios.objects.get(id_persona_id=persona.id)
                
                if not img.image:
                    if persona.sexo == 'M':
                        imgn = '/static/images/profile/default.jpg'
                    else:
                        imgn = '/static/images/profile/defaultF.png'
                else:
                    imgn = img.image

                informationConsultorUser = Consultores.objects.filter(id_persona_id=persona.id).first()
                projects = ProyectoConsultor.objects.filter(id_consultor_id=informationConsultorUser.id)

                if projects:
                    cantidad_proyectos = projects.count()
                    puntuacion = 0

                    for i in projects:
                        puntos = i.puntuacion
                        if puntos is not None:  # Verificar si el valor no es None
                            puntuacion += float(puntos)

                    if cantidad_proyectos != 0:  # Asegurarse de que no se esté dividiendo por cero
                        puntuacionConsultor = puntuacion / cantidad_proyectos
                        puntuacionConsultor_entero = int(puntuacionConsultor)
                    else:
                        puntuacionConsultor_entero = 0
                else:
                    puntuacionConsultor_entero = 0


                # Agregar los campos que desees incluir en la respuesta
                resultados.append({
                    'id':persona.id,
                    'nombre': persona.nombre,
                    'ape_pat': persona.ape_pat,
                    'ape_mat': persona.ape_mat,
                    'municipio':persona.municipio,
                    'ciudad':persona.ciudad,
                    'colonia':persona.colonia,
                    'image':imgn,
                    'puntuacionConsultor_entero':puntuacionConsultor_entero
                    # Agrega más campos si es necesario
                })
            # print(resultados)

            # Crear el diccionario de respuesta
            response_data = {
                'status': 200,  # Código de estado de respuesta exitosa
                'message': 'Búsqueda exitosa',
                'results': resultados,
            }

            return JsonResponse(response_data, safe=False)
        except Exception as e:
            print(e)
            return HttpResponse(status=500)




@login_required
def queryForName(request):
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre', '')
            apepa = request.POST.get('apepa', '')
            apema = request.POST.get('apema', '')

            personas = Personas.objects.filter(nombre__icontains=nombre, ape_pat__icontains=apepa, ape_mat__icontains=apema)

            consultores = Consultores.objects.filter(id_persona__in=personas)

            correos = []
            for consultor in consultores:
                usuarios = Usuarios.objects.filter(id_persona=consultor.id_persona, is_superuser=0)
                if usuarios.exists():
                    correo = usuarios[0].correo
                    correos.append(correo)

            usuariosPersona = Usuarios.objects.filter(correo__in=correos)
            personasResultado = []
            for usuario in usuariosPersona:
                persona = Personas.objects.get(pk=usuario.id_persona_id)
                personasResultado.append(persona)

            resultados = []
            for persona in personasResultado:
                img = Usuarios.objects.get(id_persona_id=persona.id)
                
                if not img.image:
                    if persona.sexo == 'M':
                        imgn = '/static/images/profile/default.jpg'
                    else:
                        imgn = '/static/images/profile/defaultF.png'
                else:
                    imgn = img.image


                informationConsultorUser = Consultores.objects.get(id_persona_id=persona.id)
                projects = ProyectoConsultor.objects.filter(id_consultor_id=informationConsultorUser.id)

                if projects:
                    cantidad_proyectos = projects.count()
                    puntuacion = 0

                    for i in projects:
                        puntos = i.puntuacion
                        if puntos is not None:  # Verificar si el valor no es None
                            puntuacion += float(puntos)

                    if cantidad_proyectos != 0:  # Asegurarse de que no se esté dividiendo por cero
                        puntuacionConsultor = puntuacion / cantidad_proyectos
                        puntuacionConsultor_entero = int(puntuacionConsultor)
                    else:
                        puntuacionConsultor_entero = 0
                else:
                    puntuacionConsultor_entero = 0


                # Agregar los campos que desees incluir en la respuesta
                resultados.append({
                    'id':persona.id,
                    'nombre': persona.nombre,
                    'ape_pat': persona.ape_pat,
                    'ape_mat': persona.ape_mat,
                    'municipio':persona.municipio,
                    'ciudad':persona.ciudad,
                    'colonia':persona.colonia,
                    'image':imgn,
                    'puntuacionConsultor_entero':puntuacionConsultor_entero
                    # Agrega más campos si es necesario
                })
            # print(resultados)
            # Crear el diccionario de respuesta
            response_data = {
                'status': 200,  # Código de estado de respuesta exitosa
                'message': 'Búsqueda exitosa',
                'results': resultados,
            }

            return JsonResponse(response_data, safe=False)
        except Exception as e:
            print(e)
            return HttpResponse(status=500)



@login_required
def getConsultoresDisponibles(request):
    if request.method == 'GET':
        try:
            nombre = request.POST.get('nombre', '')
            apepa = request.POST.get('apepa', '')
            apema = request.POST.get('apema', '')

            personas = Personas.objects.filter(disponible=1)

            consultores = Consultores.objects.filter(id_persona__in=personas)

            correos = []
            for consultor in consultores:
                usuarios = Usuarios.objects.filter(id_persona=consultor.id_persona, is_superuser=0)
                if usuarios.exists():
                    correo = usuarios[0].correo
                    correos.append(correo)

            usuariosPersona = Usuarios.objects.filter(correo__in=correos)
            personasResultado = []
            for usuario in usuariosPersona:
                persona = Personas.objects.get(pk=usuario.id_persona_id)
                personasResultado.append(persona)

            resultados = []
            for persona in personasResultado:
                img = Usuarios.objects.get(id_persona_id=persona.id)
                
                if not img.image:
                    if persona.sexo == 'M':
                        imgn = '/static/images/profile/default.jpg'
                    else:
                        imgn = '/static/images/profile/defaultF.png'
                else:
                    imgn = img.image


                informationConsultorUser = Consultores.objects.filter(id_persona_id=persona.id).first()
                projects = ProyectoConsultor.objects.filter(id_consultor_id=informationConsultorUser.id)

                if projects:
                    cantidad_proyectos = projects.count()
                    puntuacion = 0

                    for i in projects:
                        puntos = i.puntuacion
                        if puntos is not None:  # Verificar si el valor no es None
                            puntuacion += float(puntos)

                    if cantidad_proyectos != 0:  # Asegurarse de que no se esté dividiendo por cero
                        puntuacionConsultor = puntuacion / cantidad_proyectos
                        puntuacionConsultor_entero = int(puntuacionConsultor)
                    else:
                        puntuacionConsultor_entero = 0
                else:
                    puntuacionConsultor_entero = 0

                # Agregar los campos que desees incluir en la respuesta
                resultados.append({
                    'id':persona.id,
                    'nombre': persona.nombre,
                    'ape_pat': persona.ape_pat,
                    'ape_mat': persona.ape_mat,
                    'municipio':persona.municipio,
                    'ciudad':persona.ciudad,
                    'colonia':persona.colonia,
                    'image':imgn,
                    'puntuacionConsultor_entero':puntuacionConsultor_entero
                    # Agrega más campos si es necesario
                })

            # Crear el diccionario de respuesta
            response_data = {
                'status': 200,  # Código de estado de respuesta exitosa
                'message': 'Búsqueda exitosa',
                'results': resultados,
            }

            return JsonResponse(response_data, safe=False)
        except Exception as e:
            print(e)
            return HttpResponse(status=500)


@login_required
def consultorDocumentacion(request):
    try:
        # Lógica para obtener la ruta del archivo PDF
        # ruta_pdf = "PDF/México/456784567895/456784567895_CV.pdf"
        pais = request.GET.get('pais', None)  
        rfc = request.GET.get('rfc', None)
        id = request.GET.get('id', None)  
        format = request.GET.get('format', None)  
        file = request.GET.get('file', None)
        nombre_pdf = rfc+'_'+file+'.'+format
        # consultor = Usuarios.objects.get(pk=68)
        informationConsultorUser = Consultores.objects.get(pk=int(id))
        informationPersonalUser = Personas.objects.get(pk=informationConsultorUser.id_persona_id)

        # Generar la URL para el archivo PDF
        url_pdf = reverse('servir_pdf', kwargs={'pais': pais, 'rfc': rfc, 'nombre_pdf': nombre_pdf})
        # Redirigir al usuario a la URL del PDF
        return redirect(url_pdf)
    
    except Consultores.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('principalAdmin')
    except Personas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('principalAdmin')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')


def servir_pdf(request, pais, rfc, nombre_pdf):
    ruta_pdf = os.path.join(settings.MEDIA_ROOT, 'PDF', pais, rfc, nombre_pdf)
    if os.path.exists(ruta_pdf):
        return FileResponse(open(ruta_pdf, 'rb'), content_type='application/pdf')
    else:
        return HttpResponseNotFound()


def calcular_edad(fecha_nacimiento):
    hoy = datetime.today()
    edad = hoy.year - fecha_nacimiento.year
    if hoy.month < fecha_nacimiento.month or (hoy.month == fecha_nacimiento.month and hoy.day < fecha_nacimiento.day):
        edad -= 1
    return edad


def obtener_mes(fecha):
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    mes = fecha.month
    mes = int(mes)
    mesNumber = mes-1
    return meses[mesNumber] + ' ' + str(fecha.year)


def searchFile(ruta):
    if os.path.exists(ruta):
        return '1'
    else:
        return '0'



@login_required
def agergarAdministrador(request):
    try:
        if request.method == 'GET':
            user = Usuarios.objects.get(correo=request.session.get('username'))
            informationPersonalUser = Personas.objects.get(pk=user.id_persona_id)
            notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')
            # //Pagination
            # print(notification_list)
            # //Cantidad de notificaciones que apareceran antes de crear otra page
            if notification_list.exists():
                paginator = Paginator(notification_list, 6)
                page = request.GET.get('page')
                notification = paginator.get_page(page)

                # // CANT. NOTIFICATIONS PENDING
                pending_total= NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')

            # // ONLY 4 PENDING 
                pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')[:4]
            else:
                notification = []
                pending = []
                pending_total = []

            return render(request, 'admin/add_consultor.html',{
                'informationPersonalUser': informationPersonalUser,
                'indice':'Consultores',
                'notifications':notification,
                'pending':pending,
                'pending_total':pending_total,
            })

        else:
            if request.POST.get('telefono') != '' and request.POST.get('pais') != '' and request.POST.get('cod_post') != '' and request.POST.get('estado') != '' and request.POST.get('fecha_nacimiento') != '' and request.POST.get('genero') != '' and request.POST.get('disponibilidad') != '' and request.POST.get('nombre') != '' and request.POST.get('ape_pat') != '' and request.POST.get('correo') != '' and request.POST.get('password1') != '' and request.POST.get('password2') != '' and request.POST.get('reference') != '':
                
                if request.POST.get('password1') == request.POST.get('password2'):
                    name_value = request.POST['nombre']
                    apepa_value = request.POST['ape_pat']
                    apema_value = request.POST['ape_mat']
                    correo_value = request.POST['correo']
                    password_value = request.POST['password1']
                    
                    try:
                        query = Usuarios.objects.get(correo=correo_value)
                        messages.error(
                        request, 'El correo ya esta registrado para otra cuenta, utilice otro por favor')
                        return redirect('agergarAdministrador') 

                    except Usuarios.DoesNotExist:
                        params = {
                            'name':  cipher_suite.encrypt(name_value.encode('utf-8')).decode('utf-8'),
                            'apepa': cipher_suite.encrypt(apepa_value.encode('utf-8')).decode('utf-8'),
                            'apema': cipher_suite.encrypt(apema_value.encode('utf-8')).decode('utf-8'),
                            'email': cipher_suite.encrypt(correo_value.encode('utf-8')).decode('utf-8'),
                            'password': cipher_suite.encrypt(password_value.encode('utf-8')).decode('utf-8'),
                            'telefono':  cipher_suite.encrypt(request.POST.get('telefono').encode('utf-8')).decode('utf-8'),
                            'pais': cipher_suite.encrypt(request.POST.get('pais').encode('utf-8')).decode('utf-8'),
                            'cod_post': cipher_suite.encrypt(request.POST.get('cod_post').encode('utf-8')).decode('utf-8'),
                            'estado': cipher_suite.encrypt(request.POST.get('estado').encode('utf-8')).decode('utf-8'),
                            'ciudad': cipher_suite.encrypt(request.POST.get('ciudad').encode('utf-8')).decode('utf-8'),
                            'municipio': cipher_suite.encrypt(request.POST.get('municipio').encode('utf-8')).decode('utf-8'),
                            'colonia': cipher_suite.encrypt(request.POST.get('colonia').encode('utf-8')).decode('utf-8'),
                            'fecha_nacimiento': cipher_suite.encrypt(request.POST.get('fecha_nacimiento').encode('utf-8')).decode('utf-8'),
                            'genero': cipher_suite.encrypt(request.POST.get('genero').encode('utf-8')).decode('utf-8'),
                            'disponibilidad': cipher_suite.encrypt(request.POST.get('disponibilidad').encode('utf-8')).decode('utf-8'),
                            'referencia': cipher_suite.encrypt(request.POST.get('reference').encode('utf-8')).decode('utf-8'),
                        }

                        query_string = urlencode(params)
                        url = reverse('profesionConsultorAdmin') + '?'+ query_string
                        return redirect(url)
                       
                else:
                    messages.error(
                    request, 'Las contraseñas no coinciden')
                    return redirect('agergarAdministrador')

            else:
                messages.error(
                    request, 'Debes llenar todos los campos')
                return redirect('agergarAdministrador')

    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Personas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')




@login_required
def profesionConsultorAdmin(request):
    try:
        if request.method == 'GET':
            user = Usuarios.objects.get(correo=request.session.get('username'))
            informationPersonalUser = Personas.objects.get(pk=user.id_persona_id)
            notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')
            

            if notification_list.exists():
                paginator = Paginator(notification_list, 6)
                page = request.GET.get('page')
                notification = paginator.get_page(page)

                
                pending_total= NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')

            
                pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')[:4]
            else:
                notification = []
                pending = []
                pending_total = []

            monedaCobro = list(TipoMoneda.objects.values())
            maneraPago = list(ManeraPago.objects.values())
            niveles = list(NivelesConocimiento.objects.values())


            return render(request, 'admin/profesion_consultor.html',{
                'informationPersonalUser': informationPersonalUser,
                'indice':'Consultores',
                'monedaCobro':monedaCobro,
                'maneraPago': maneraPago,
                'niveles':niveles,
                'notifications':notification,
                'pending':pending,
                'pending_total':pending_total,
            })

        else:
            if request.POST.get('puesto') != '' and request.POST.get('tipo_moneda') != '' and request.POST.get('tarifa') != '' and request.POST.get('forma_cobro') != '' and request.POST.get('RFC') != '' and 'accepted_emails' in request.POST and request.POST.get('SAP') != '' and request.POST.get('expriencia') != '':
                carga = []
                params = {
                    'puesto':  cipher_suite.encrypt(request.POST.get('puesto').encode('utf-8')).decode('utf-8'),
                    'tipo_moneda':  cipher_suite.encrypt(request.POST.get('tipo_moneda').encode('utf-8')).decode('utf-8'),
                    'tarifa':  cipher_suite.encrypt(request.POST.get('tarifa').encode('utf-8')).decode('utf-8'),
                    'forma_cobro':  cipher_suite.encrypt(request.POST.get('forma_cobro').encode('utf-8')).decode('utf-8'),
                    'RFC':  cipher_suite.encrypt(request.POST.get('RFC').encode('utf-8')).decode('utf-8'),
                    'SAP':  cipher_suite.encrypt(request.POST.get('SAP').encode('utf-8')).decode('utf-8'),
                    'experiencia':  cipher_suite.encrypt(request.POST.get('experiencia').encode('utf-8')).decode('utf-8'),
                }

                query_string = urlencode(params)
                url = reverse('experienceConsultorAdmin') + '?' + request.GET.urlencode() + '&' + query_string
                return redirect(url)

            else:
                messages.error(
                    request, 'Debes llenar todos los campos')
                return redirect('agergarAdministrador')

    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Personas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')




@login_required
def experienceConsultorAdmin(request):
    try:
        if request.method == 'GET':
            user = Usuarios.objects.get(correo=request.session.get('username'))
            informationPersonalUser = Personas.objects.get(pk=user.id_persona_id)
            notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')
            encrypted_fecha_nacimiento = request.GET.get('fecha_nacimiento').encode('utf-8')
            fechaNac = cipher_suite.decrypt(encrypted_fecha_nacimiento).decode('utf-8')

            if notification_list.exists():
                paginator = Paginator(notification_list, 6)
                page = request.GET.get('page')
                notification = paginator.get_page(page)                
                pending_total= NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')
                pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')[:4]
            else:
                notification = []
                pending = []
                pending_total = []
            encrypted_fecha_nacimiento = request.GET.get('fecha_nacimiento').encode('utf-8')
            fechaNac = cipher_suite.decrypt(encrypted_fecha_nacimiento).decode('utf-8')

            return render(request, 'admin/experience_consultor.html',{
                'informationPersonalUser': informationPersonalUser,
                'indice':'Consultores',
                'fechaNac':fechaNac,
                'notifications':notification,
                'pending':pending,
                'pending_total':pending_total,
            })

        else:
            if request.POST.get('puesto') != '' and request.POST.get('empresa') != '' and request.POST.get('fecha_entrada') != '' and request.POST.get('activities') != '' and request.POST.get('notas') != '':
                notas = request.POST['notas']
                experiencia = Experiencias(nombre=request.POST['puesto'])
                experiencia.save()

                fecha_entrada_str = request.POST['fecha_entrada']
                fecha_entrada = datetime.strptime(
                    fecha_entrada_str, '%Y-%m-%d')
                fecha_salida = date.today()
                if 'chec' in request.POST:
                    tiempoExperiencia = 'Sigue Trabajando'
                else:
                    fecha_salida_str = request.POST['fecha_salida']
                    if fecha_salida_str:
                        fecha_salida = datetime.strptime(
                            fecha_salida_str, '%Y-%m-%d')
                        diferencia = fecha_salida - fecha_entrada
                        tiempoExperiencia = diferencia.days
                    else:
                        messages.error(request, 'Proporcione la fecha de salida')
                        return redirect('agergarAdministrador')
                        

                experienciaConsultor = ExperienciasConsultor(id_experiencia=experiencia,empresa=request.POST['empresa'], puesto=request.POST['puesto'], descripcion=request.POST['activities'], tiempo_experiencia=tiempoExperiencia, fecha_entrada=fecha_entrada, fecha_salida=fecha_salida)
                experienciaConsultor.save()

                data = [notas, experiencia.id, experienciaConsultor.id] 

                parametros_get = {k: v for k, v in request.GET.items() if k.startswith('cargaExperiencia')}
                contador = len(parametros_get)

                contador += 1
                numCarga = f'cargaExperiencia_{contador}'

                # Crear un nuevo diccionario con la clave-valor actual
                nuevo_parametro = {numCarga: cipher_suite.encrypt(str(data).encode('utf-8')).decode('utf-8')}

                # Definir el diccionario params y actualizarlo con el nuevo parámetro
                params = {k: v for k, v in request.GET.items()}
                params.update(nuevo_parametro)

                # for key, value in request.POST.items():
                    # print(f'{key}: {value}')

                if request.POST.get('addMore') == 'ON':
                    url = reverse('experienceConsultorAdmin') + '?' + urlencode(params)
                    return redirect(url)

                else:
                    query_string = urlencode(params)
                    url = reverse('educationConsultorAdmin') + '?' + query_string
                    return redirect(url)


            else:
                messages.error(
                    request, 'Debes llenar todos los campos')
                return redirect('profesionConsultorAdmin')

    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Personas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')





@login_required
def educationConsultorAdmin(request):
    try:
        if request.method == 'GET':
            user = Usuarios.objects.get(correo=request.session.get('username'))
            informationPersonalUser = Personas.objects.get(pk=user.id_persona_id)
            notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')
            
            if notification_list.exists():
                paginator = Paginator(notification_list, 6)
                page = request.GET.get('page')
                notification = paginator.get_page(page)
                pending_total= NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')
                pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')[:4]
            else:
                notification = []
                pending = []
                pending_total = []

            encrypted_fecha_nacimiento = request.GET.get('fecha_nacimiento').encode('utf-8')
            fechaNac = cipher_suite.decrypt(encrypted_fecha_nacimiento).decode('utf-8')
            
            return render(request, 'admin/education_consultor.html',{
                'informationPersonalUser': informationPersonalUser,
                'indice':'Consultores',
                'fechaNac':fechaNac,
                'notifications':notification,
                'pending':pending,
                'pending_total':pending_total,
            })

        else:
            if request.POST.get('nivel') != '' and request.POST.get('institucion') != '' and request.POST.get('titulo') != '' and request.POST.get('ano_inicio') != '':
                
                if Instituciones.objects.filter(nombre=request.POST['institucion']).exists():
                    # Se encontró la institución
                    institucion = Instituciones.objects.get(nombre=request.POST['institucion'])
                else:
                    # No se encontró la institución
                    institucion = Instituciones(nombre=request.POST['institucion'])
                
                    institucion.save()

                if 'chec' in request.POST:
                    estudios = Estudios(titulo_registrado=request.POST['titulo'], fecha_ingreso=request.POST['ano_inicio'], fecha_termino='Estudiando', educacion=request.POST['nivel'], id_institucion=institucion)
                else:
                    estudios = Estudios(titulo_registrado=request.POST['titulo'], fecha_ingreso=request.POST['ano_inicio'], fecha_termino=request.POST['ano_termino'], educacion=request.POST['nivel'], id_institucion=institucion)
                estudios.save()
                
                educacion = [institucion.id, estudios.id]

                parametros_get = {k: v for k, v in request.GET.items() if k.startswith('cargaEstudios')}
                contador = len(parametros_get)

                contador += 1
                numCarga = f'cargaEstudios_{contador}'

                # Crear un nuevo diccionario con la clave-valor actual
                nuevo_parametro = {numCarga: cipher_suite.encrypt(str(educacion).encode('utf-8')).decode('utf-8')}

                # Definir el diccionario params y actualizarlo con el nuevo parámetro
                params = {k: v for k, v in request.GET.items()}
                params.update(nuevo_parametro)

                    
                if request.POST.get('addMore') == 'ON':
                    url = reverse('educationConsultorAdmin') + '?' + urlencode(params)
                    return redirect(url)

                else:
                    query_string = urlencode(params)
                    url = reverse('AgregadoConsultorAdmin') + '?' + query_string
                    return redirect(url)

            else:
                messages.error(
                    request, 'Debes llenar todos los campos')
                return redirect('agergarAdministrador')

    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Personas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')







@login_required
def agregadoConsultorAdmin(request):
    try:
        if request.method == 'GET':
            user = Usuarios.objects.get(correo=request.session.get('username'))
            informationPersonalUser = Personas.objects.get(pk=user.id_persona_id)
            notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')
            
            if notification_list.exists():
                paginator = Paginator(notification_list, 6)
                page = request.GET.get('page')
                notification = paginator.get_page(page)
                pending_total= NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')
                pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')[:4]
            else:
                notification = []
                pending = []
                pending_total = []

            encrypted_nombre = request.GET.get('name').encode('utf-8') #si
            nombre = cipher_suite.decrypt(encrypted_nombre).decode('utf-8') #si
            encrypted_apepa = request.GET.get('apepa').encode('utf-8') #si
            apepa = cipher_suite.decrypt(encrypted_apepa).decode('utf-8') #si
            encrypted_apema = request.GET.get('apema').encode('utf-8') #si
            apema = cipher_suite.decrypt(encrypted_apema).decode('utf-8') #si
            encrypted_email = request.GET.get('email').encode('utf-8') #si
            email = cipher_suite.decrypt(encrypted_email).decode('utf-8') #si
            encrypted_password = request.GET.get('password').encode('utf-8') #si
            password = cipher_suite.decrypt(encrypted_password).decode('utf-8') #si
            encrypted_telefono = request.GET.get('telefono').encode('utf-8') #si
            telefono = cipher_suite.decrypt(encrypted_telefono).decode('utf-8') #si
            encrypted_pais = request.GET.get('pais').encode('utf-8') #si
            pais = cipher_suite.decrypt(encrypted_pais).decode('utf-8') #si
            encrypted_cod_post = request.GET.get('cod_post').encode('utf-8') #si
            cod_post = cipher_suite.decrypt(encrypted_cod_post).decode('utf') #si
            encrypted_estado = request.GET.get('estado').encode('utf-8') #si
            estado = cipher_suite.decrypt(encrypted_estado).decode('utf') #si
            encrypted_ciudad = request.GET.get('ciudad').encode('utf-8') #si
            ciudad = cipher_suite.decrypt(encrypted_ciudad).decode('utf') #si
            encrypted_municipio = request.GET.get('municipio').encode('utf-8') #si
            municipio = cipher_suite.decrypt(encrypted_municipio).decode('utf') #si
            encrypted_colonia = request.GET.get('colonia').encode('utf-8') #si
            colonia = cipher_suite.decrypt(encrypted_colonia).decode('utf') #si
            encrypted_fecha_nacimiento = request.GET.get('fecha_nacimiento').encode('utf-8') #si
            fecha_nacimiento = cipher_suite.decrypt(encrypted_fecha_nacimiento).decode('utf') #si
            encrypted_genero = request.GET.get('genero').encode('utf-8') #si
            genero = cipher_suite.decrypt(encrypted_genero).decode('utf') #si
            encrypted_disponibilidad = request.GET.get('disponibilidad').encode('utf-8') #si
            disponibilidad = cipher_suite.decrypt(encrypted_disponibilidad).decode('utf') #si
            

            encrypted_SAP = request.GET.get('SAP').encode('utf-8') #si
            SAP = cipher_suite.decrypt(encrypted_SAP).decode('utf') #si
            encrypted_referencia = request.GET.get('referencia').encode('utf-8') #si
            referencia = cipher_suite.decrypt(encrypted_referencia).decode('utf') #si
            encrypted_puesto = request.GET.get('puesto').encode('utf-8') # si
            puesto = cipher_suite.decrypt(encrypted_puesto).decode('utf') # si
            encrypted_tipo_moneda = request.GET.get('tipo_moneda').encode('utf-8') #si
            tipo_moneda = cipher_suite.decrypt(encrypted_tipo_moneda).decode('utf') #si
            encrypted_tarifa = request.GET.get('tarifa').encode('utf-8') #si
            tarifa = cipher_suite.decrypt(encrypted_tarifa).decode('utf') #si
            encrypted_forma_cobro = request.GET.get('forma_cobro').encode('utf-8') #si
            forma_cobro = cipher_suite.decrypt(encrypted_forma_cobro).decode('utf') #si
            encrypted_RFC = request.GET.get('RFC').encode('utf-8') #si
            RFC = cipher_suite.decrypt(encrypted_RFC).decode('utf') #si
            encrypted_SAP = request.GET.get('SAP').encode('utf-8') #si
            SAP = cipher_suite.decrypt(encrypted_SAP).decode('utf') #si

            encrypted_experiencia = request.GET.get('experiencia').encode('utf-8')
            experiencia = cipher_suite.decrypt(encrypted_experiencia).decode('utf') 


            parametros_get = {k: v for k, v in request.GET.items() if k.startswith('cargaExperiencia')}

            valores_carga_experiencia = []

            for key, value in parametros_get.items():
                experiencia_descifrada = cipher_suite.decrypt(value.encode('utf-8')).decode('utf-8')
                valores_carga_experiencia.append(experiencia_descifrada)

        

            parametros_get = {k: v for k, v in request.GET.items() if k.startswith('cargaEstudios')}

            valores_carga_estudios = []

            for key, value in parametros_get.items():
                experiencia_descifrada = cipher_suite.decrypt(value.encode('utf-8')).decode('utf-8')
                valores_carga_estudios.append(experiencia_descifrada)

        


            persona = Personas(nombre=nombre, ape_pat=apepa, ape_mat=apema, fecha_nacimiento=fecha_nacimiento, ciudad=ciudad, cod_post=cod_post, estado=estado, pais=pais, telefono=telefono, sexo=genero, municipio=municipio, colonia=colonia, disponible=int(disponibilidad), referencia=referencia)
            persona.save()

            # print(persona)
            
            user = Usuarios.objects.create_superuser(correo=email, password=password, id_persona=persona, rol='Administrador')
            user.save()
            # print(user)

            pago = ManeraPago.objects.get(pk=int(forma_cobro))
            moneda = TipoMoneda.objects.get(pk=int(tipo_moneda))
            nivel = NivelesConocimiento.objects.get(pk=int(experiencia))

            consultores = Consultores(
                    tipo_persona='Fisica', rfc=RFC, tarifa_hora=int(tarifa), id_persona=persona, id_manera_pago=pago, id_tipo_moneda=moneda, especialidad=SAP, id_nivel=nivel)
            consultores.save()
            # educacion = [institucion, estudios]
            # data = [notas, experiencia, experienciaConsultor] 

            
            for valores_experiencia in valores_carga_estudios:
                lista_real = ast.literal_eval(valores_experiencia)
                numero1 = lista_real[0]
                
                numero2 = lista_real[1]
                estudios = Estudios.objects.get(pk=int(numero2))
                estudios.id_consultor = consultores
                estudios.save()

            
            for valores_experiencia in valores_carga_experiencia:
                lista_real = ast.literal_eval(valores_experiencia)
                # Acceder a los elementos individuales
                cadena_texto = lista_real[0]
                nota = NotasGnosisConsultor(nota=cadena_texto, id_consultor=consultores)
                nota.save()

                numero1 = lista_real[1]
                numero2 = lista_real[2]
                experiencia = ExperienciasConsultor.objects.get(pk=int(numero2))
                experiencia.id_consultor = consultores
                experiencia.save()
            

            return render(request, 'admin/count_created.html',{
                'informationPersonalUser': informationPersonalUser,
                'indice':'Consultores',
                'user':user.correo,
                'notifications':notification,
                'pending':pending,
                'pending_total':pending_total,
            })

        
    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Personas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')






@login_required
def proyectosForAdmin(request):
    try:
        user = Usuarios.objects.get(correo=request.session.get('username'))
        informationPersonalUser = Personas.objects.get(pk=user.id_persona_id)
        
        consulta = request.GET.get('consulta')
        estado = request.GET.get('estado')
        ciudad = request.GET.get('ciudad')

        experiencia = request.GET.get('experiencia')
        tipo = request.GET.get('tipo')

        registros = Proyectos.objects.all().order_by('-id')

        if consulta:
            registros = registros.filter(proyecto_nombre__icontains=consulta)

        if estado:
            empresas = Empresas.objects.filter(estado__icontains=estado)
            empresaProyecto = EmpresaProyecto.objects.filter(id_empresa__in=empresas)
            registros = registros.filter(id_empresa_proyecto__in=empresaProyecto)

        if ciudad:
            empresas = Empresas.objects.filter(ciudad__icontains=ciudad)
            empresaProyecto = EmpresaProyecto.objects.filter(id_empresa__in=empresas)
            registros = registros.filter(id_empresa_proyecto__in=empresaProyecto)


        if tipo:
            registros = registros.filter(tipo__exact=tipo)

        if estado and ciudad:
            empresas = Empresas.objects.filter(estado__icontains=estado, ciudad__icontains=ciudad)
            empresaProyecto = EmpresaProyecto.objects.filter(id_empresa__in=empresas)
            registros = registros.filter(id_empresa_proyecto__in=empresaProyecto)        
        
        requerimientos = RequerimientosModulosProyecto.objects.all()
        
        if experiencia:
            nivelesConocimiento = NivelesConocimiento.objects.filter(nombre__exact=experiencia)
            requerimientos_ids = RequerimientosModulosProyecto.objects.filter(id_experiencia_requerida__in=nivelesConocimiento).values_list('id_proyecto', flat=True)

            # print(requerimientos_ids)
            registros = registros.filter(id__in=requerimientos_ids)


        paginator = Paginator(registros, 8)
        page_number = request.GET.get('page')
        proyectos = paginator.get_page(page_number)
        cantidad_proyectos = paginator.count
        num_paginas = paginator.num_pages
        proyectos_con_requerimientos = {}

        for proyecto in proyectos:
            proyectos_con_requerimientos[proyecto] = []
            requerimientos_proyecto = requerimientos.filter(id_proyecto=proyecto)
            proyectos_con_requerimientos[proyecto] = requerimientos_proyecto

        
        notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')
        # //Pagination
        
        # //Cantidad de notificaciones que apareceran antes de crear otra page
        if notification_list.exists():
            paginator = Paginator(notification_list, 6)
            page = request.GET.get('page')
            notification = paginator.get_page(page)
            # // CANT. NOTIFICATIONS PENDING
            pending_total= NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')
            # // ONLY 4 PENDING 
            pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')[:4]
        else:
            notification = []
            pending = []
            pending_total = []
        listaIdiomas = list(Idiomas.objects.values())
        listaModulos = list(Modulos.objects.values())
        listaSubodulos = list(Submodulos.objects.values())
        listaNivelesConocimiento = list(NivelesConocimiento.objects.values())
        return render(request, 'proyectos/verProyectos.html',{
            'indice':'Proyectos',
            'image':user.image,
            'proyectos_con_requerimientos': proyectos_con_requerimientos,
            'num_paginas':num_paginas,
            'informationPersonalUser': informationPersonalUser,
            'cantidad_proyectos':cantidad_proyectos,
            'notifications':notification,
            'pending':pending,
            'pending_total':pending_total,
            'listaIdiomas': listaIdiomas,
            'listaModulos':listaModulos,
            'listaSubodulos':listaSubodulos,
            'listaNivelesConocimiento':listaNivelesConocimiento,
        })

    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Personas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')



@login_required
def detallesProyectoAdmin(request, id=int):
    try:
        user = Usuarios.objects.get(correo=request.session.get('username'))
        informationPersonalUser = Personas.objects.get(pk=user.id_persona_id)
        
        proyecto = Proyectos.objects.get(pk=id)
        empresaProyecto = EmpresaProyecto.objects.filter(id_proyecto=proyecto).first()
        if empresaProyecto is not None:
            empresa = Empresas.objects.get(pk=empresaProyecto.id_empresa_id)
        else:
            empresa = 'Desconocida'


        requerimientosModulosProyecto = RequerimientosModulosProyecto.objects.filter(id_proyecto=proyecto).order_by('-id')
        requerimientosIdiomasProyecto = RequerimientosIdiomasProyecto.objects.filter(id_proyecto=proyecto).order_by('-id')

        modulos_ids = requerimientosModulosProyecto.values_list('id_modulo', flat=True)
        idiomas_ids = requerimientosIdiomasProyecto.values_list('id_idioma', flat=True)

        personas = Personas.objects.filter(disponible=True)
        consultores = Consultores.objects.filter(id_persona__in=personas)

        # Usamos Q para construir la consulta de manera más flexible
        consulta_modulos = Q(conocimientosconsultor__id_modulo__in=modulos_ids) if modulos_ids else Q()
        consulta_idiomas = Q(idiomasconsultor__id_idioma__in=idiomas_ids) if idiomas_ids else Q()

        consultores_cumplen_requerimientos = consultores.filter(consulta_modulos | consulta_idiomas)
        consultores_cumplen_requerimientos = consultores_cumplen_requerimientos.order_by('-id_categoria__id')

        dataConsultores = []
        for data in consultores_cumplen_requerimientos:
            con = data.id
            id = data.id_persona.id
            if data.id_persona.sexo == 'M':
                sexo = 'hombre'
            else:
                sexo = 'Mujer'  

            if data.id_persona.disponible:
                disponible = 'Disponible'
            else:
                disponible = 'No disponible'

            edad = calcular_edad(str(data.id_persona.fecha_nacimiento))


            modulos = ConocimientosConsultor.objects.filter(id_consultor=data)
            
            experiencia = ""
            if modulos:
                for modulo in modulos:
                    experiencia += modulo.id_modulo.nombre + " "
            else:
                experiencia = "Sin experiencia"
                    

            if hasattr(data, 'tarifa_hora'):
                pago = data.tarifa_hora
            else:
                pago = None

            comentarios = 'Aun no hay comentarios'


            if hasattr(data.id_tipo_moneda, 'tipo'):
                texto = data.id_tipo_moneda.tipo
                resultado = re.search(r'\((.*?)\)', texto)
                moneda = resultado.group(1)
            else:
                moneda = None

            projects = ProyectoConsultor.objects.filter(id_consultor_id=data.id)

            if projects:
                cantidad_proyectos = projects.count()
                puntuacion = 0

                for i in projects:
                    puntos = i.puntuacion
                    if puntos is not None:  # Verificar si el valor no es None
                         puntuacion += float(puntos)

                if cantidad_proyectos != 0:  # Asegurarse de que no se esté dividiendo por cero
                    puntuacionConsultor = puntuacion / cantidad_proyectos
                    puntuacionConsultor_entero = int(puntuacionConsultor)
                else:
                    puntuacionConsultor_entero = 0
            else:
                puntuacionConsultor_entero = 0



            dataConsultores.append([sexo, edad, experiencia, pago, disponible, moneda, comentarios, id, con, puntuacionConsultor_entero])
        


        fecha_actual = datetime.now().date()
        # Fecha de publicación del staffing (ejemplo)
        fecha_publicacion = proyecto.fecha_publicacion

        # Calcula la diferencia de días
        diferencia = fecha_actual - fecha_publicacion
        dias_pasados = diferencia.days

        if dias_pasados > 30:
            diasStaffing = 'Hace mas de un mes'
        elif dias_pasados == 0:
            diasStaffing = 'Hoy'
        else:
            diasStaffing = 'Hace ' + str(dias_pasados) + ' días'

        notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')
        # //Pagination
        
        # //Cantidad de notificaciones que apareceran antes de crear otra page
        if notification_list.exists():
            paginator = Paginator(notification_list, 6)
            page = request.GET.get('page')
            notification = paginator.get_page(page)
            # // CANT. NOTIFICATIONS PENDING
            pending_total= NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')
            # // ONLY 4 PENDING 
            pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalUser.id).order_by('-created_at')[:4]
        else:
            notification = []
            pending = []
            pending_total = []

        return render(request, "proyectos/detalles_proyecto.html", {
            'indice':'Proyectos',
            'image':user.image,
            'proyecto':proyecto,
            'dataConsultores':dataConsultores,
            'empresa':empresa,
            'requerimientosModulosProyecto':requerimientosModulosProyecto,
            'diasStaffing':diasStaffing,
            'informationPersonalUser': informationPersonalUser,
            'requerimientosIdiomasProyecto':requerimientosIdiomasProyecto,
            'notifications':notification,
            'pending':pending,
            'pending_total':pending_total,
        })
        
    except Consultores.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Personas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')





@login_required
def postularByAdministrador(request):
    try:
        data = json.loads(request.body)
        id = data['proyecto']
        consultor = data['consultor']
        
        informationConsultorUser = Consultores.objects.get(pk=int(consultor))
        informationPersonalUser = Personas.objects.get(pk=informationConsultorUser.id_persona_id)

        proyecto = Proyectos.objects.get(pk=int(id))
        empresa = proyecto.id_empresa_proyecto.id_empresa

        postulacion = PostulacionesProyectoGnosis.objects.filter(id_proyecto=proyecto, id_empresa=empresa, id_consultor=informationConsultorUser)

        if postulacion:
            pass
        else:
            postulacion = PostulacionesProyectoGnosis(id_proyecto=proyecto, id_empresa=empresa, id_consultor=informationConsultorUser)
            postulacion.save()
            
            
            name = 'Gnosis SC'
            
            if informationPersonalUser.sexo == 'M':
                titulo = '¡¡Un nuevo consultor!!'
                message = 'Hemos postulado a un nuevo consultor para su proyecto ' + str(proyecto.proyecto_nombre)
            else:
                titulo = '¡¡Una nueva consultora!!'
                message = 'Hemos postulado a una nueva consultora para su proyecto ' + str(proyecto.proyecto_nombre)

            notificacion = NotificationEmpresa(name=name, email='gnosis@gnosis.com.mx', subject=titulo, message=message, status='Pending', id_empresa_destinatary=empresa)
            notificacion.save()
        
        return HttpResponse(status=200)
    except Consultores.DoesNotExist as error:
        print(f"Error: {str(error)}")
        return HttpResponse(status=500)

    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        return HttpResponse(status=500)
    except Personas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        return HttpResponse(status=500)
    except Exception as error:
        print(f"Error: {str(error)}")
        return HttpResponse(status=500)



def calcular_edad(fecha_nacimiento):
    fecha_actual = date.today()
    fecha_nacimiento = datetime.strptime(fecha_nacimiento, "%Y-%m-%d %H:%M:%S%z").date()
    edad = fecha_actual.year - fecha_nacimiento.year

    # Comprobar si aún no ha llegado el cumpleaños de este año
    if fecha_actual.month < fecha_nacimiento.month or (fecha_actual.month == fecha_nacimiento.month and fecha_actual.day < fecha_nacimiento.day):
        edad -= 1

    return edad




@login_required
def updatePuntuacion(request, id, prj):
    try:
        user = Usuarios.objects.get(correo=request.session.get('username'))
        informationPersonalAdministradorUser = Personas.objects.get(pk=user.id_persona_id)
    
        notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')
        # //Pagination
        # //Cantidad de notificaciones que apareceran antes de crear otra page
        if notification_list.exists():
            paginator = Paginator(notification_list, 6)
            page = request.GET.get('page')
            notification = paginator.get_page(page)

            # // CANT. NOTIFICATIONS PENDING
            pending_total= NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')

        # // ONLY 4 PENDING 
            pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')[:4]
        else:
            notification = []
            pending = []
            pending_total = []

        context = {
            'indice': 'Proyectos',
            'consultor':id,
            'proyecto':prj,
            'pending':pending,
            'pending_total':pending_total,
        }

        return render(request, 'proyectos/form_consultor_update_puntuacion.html', context)


    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')



@login_required
def formComentariosUpdateAdmin(request):
    try:
        user = Usuarios.objects.get(correo=request.session.get('username'))

        comentarios = request.POST['comentarios']
        calificacion = request.POST['calificacion']
        consultor = request.POST['consultor']
        proyecto = request.POST['proyecto']

        proyectoQuery = Proyectos.objects.get(pk=int(proyecto))
        consultorQuery = Consultores.objects.get(pk=int(consultor))

        query = ProyectoConsultor.objects.get(id_proyecto=proyectoQuery, id_consultor=consultorQuery)

        if query:
            query.comentario = comentarios
            query.puntuacion = calificacion
            query.participacion = False
            query.save()            

        return redirect('miConsultorProfile', id=consultorQuery.id_persona.id)


    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')



@login_required
def addCategoriaConsultor(request):
    try:
        consultor = request.POST['consultor']
        categoria = request.POST['categoria']
        
        query = Consultores.objects.get(pk=int(consultor))
        if query:
            cat = CategoriasConsultor.objects.get(pk=int(categoria))
            query.id_categoria = cat
            query.save()
        
        return redirect('miConsultorProfile', id=query.id_persona.id)

    except Exception as error:
        print(f"Error: {str(error)}")
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')


@login_required
def contratoConsultor(request, id:int, proyecto:int):
    try:
        proyectoConsultor = ProyectoConsultor.objects.get(id_proyecto_id=int(proyecto), id_consultor_id=int(id))

        try:
            contrato = Contratos.objects.get(id_proyecto_consultor=proyectoConsultor, titulo__startswith='CONTRATO MARCO')
        except Contratos.DoesNotExist:
            contrato = None

        name = Consultores.objects.get(pk=int(id))
        user = Usuarios.objects.get(correo=request.session.get('username'))
        informationPersonalUser = Personas.objects.get(pk=user.id_persona_id)
        informationPersonalAdministradorUser = Personas.objects.get(pk=user.id_persona_id)


        notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')
        # //Pagination
        # //Cantidad de notificaciones que apareceran antes de crear otra page
        if notification_list.exists():
            paginator = Paginator(notification_list, 6)
            page = request.GET.get('page')
            notification = paginator.get_page(page)

            # // CANT. NOTIFICATIONS PENDING
            pending_total= NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')

        # // ONLY 4 PENDING 
            pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')[:4]
        else:
            notification = []
            pending = []
            pending_total = []


        monedaCobro = list(TipoMoneda.objects.values())
        return render(request, 'proyectos/contratar.html',{
            'indice': 'Proyectos',
            'consultor':id,
            'name':name.id_persona.nombre,
            'apepa':name.id_persona.ape_pat,
            'apema':name.id_persona.ape_mat,
            'proyecto':proyecto,
            'monedaCobro':monedaCobro,
            'cambioHoyMXN':tipoCambioMXN(),
            'cambioHoyUSD':tipoCambioUSD(),
            'contrato':contrato,
            'proyectoConsultor':proyectoConsultor.id,
            'notifications':notification,
            'pending':pending,
            'pending_total':pending_total,
        })
        
    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')




@login_required
def contratoMachote(request):
    try:
        consultor = request.POST.get('consultor')
        proyecto = request.POST.get('proyecto')
        proyectoConsultor = request.POST.get('proyectoConsultor')
        titulo = request.POST.get('titulo')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_final = request.POST.get('fecha_final')
        fecha_firma = request.POST.get('fecha_firma')
        fecha_promesa = request.POST.get('fecha_promesa')
        tarifa = request.POST.get('tarifa')
        moneda = request.POST.get('moneda')
        puesto = request.POST.get('puesto')
        tarifaCliente = request.POST.get('tarifaCliente')

        empresaCopia = request.POST.get('empresaCopia')
        consultorCopia = request.POST.get('consultorCopia')

        # Asignar un valor por defecto si la variable es None
        empresaCopia = empresaCopia if empresaCopia is not None else '0'
        consultorCopia = consultorCopia if consultorCopia is not None else '0'

        manual = request.POST.get('manual-base')
        cambioManual = request.POST.get('tipoCambioHoy-Manual-base')
        cambioAuto = request.POST.get('tipoCambioHoy-Auto-base')
        tipoCambio = request.POST.get('cambio-base')

        facturaDia = request.POST.get('factura-dia-base')
        
        if facturaDia == 'ON':
            # print("factura dia")
            aplicada = 'Dia - Factura'
            cambioAuto = 0.00
            cambioManual = 0.00
            tipoCambio = 0.00
        else:

            if manual == '1':
                # print("manual")
                aplicada = 'Manual'
                
            else:
                # print("api")
                aplicada = 'Automatica'
            


        monedaValor = TipoMoneda.objects.get(pk=1)
        fechaInicio = formatearFecha(fecha_inicio)
        fechaFinal = formatearFecha(fecha_final)
        fechaFirma = formatearFecha(fecha_firma)
        
        cantidad = "$"+tarifa
        tarifa_diaria = convertir_a_letras(cantidad)
        vigencia = tiempoVigencia(request.POST.get('fecha_inicio'), request.POST.get('fecha_final'))

        try:
            proyectoConsultor = ProyectoConsultor.objects.get(pk=int(proyectoConsultor), id_proyecto_id=int(proyecto), id_consultor_id=int(consultor))

            documento, created = TipoContrato.objects.get_or_create(nombre=titulo)
            
            contratacion, val = TipoContrataciones.objects.get_or_create(contratacion='Nómina')
            
            try:
            
                contrato = Contratos.objects.get(id_proyecto_consultor=proyectoConsultor)
                contrato.delete()

                facturas = Facturas.objects.filter(id_proyecto_consultor=proyectoConsultor)
                if facturas:
                    for fac in facturas:
                        fac.delete()

                """contrato.fecha_firma=fecha_firma
                contrato.fecha_inicio=fecha_inicio
                contrato.fecha_fin=fecha_final
                contrato.tarifa_dia_consultor=tarifa
                contrato.tarifa_dia_cliente=tarifaCliente
                contrato.tiempo_vigencia=vigencia
                contrato.id_tipo_moneda=monedaValor
                contrato.id_tipo_contrato=documento
                contrato.id_tipo_contratacion=contratacion
                contrato.fecha_promesa=fecha_promesa
                contrato.tipoCambio=tipoCambio
                contrato.tipoCambioManual=cambioManual
                contrato.tipoCambioAuto=cambioAuto
                contrato.aplicada=aplicada
                contrato.save()"""

            except Contratos.DoesNotExist:
                pass

            contrato = Contratos(
                titulo='CONTRATO BASE',
                fecha_firma=fecha_firma,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_final,
                fecha_promesa=fecha_promesa,
                tarifa_dia_consultor=tarifa,
                tarifa_dia_cliente=tarifaCliente,
                tiempo_vigencia=vigencia,
                tipoCambio=tipoCambio,
                tipoCambioManual=cambioManual,
                tipoCambioAuto=cambioAuto,
                aplicada=aplicada,
                id_tipo_moneda=monedaValor,
                id_tipo_contratacion=contratacion,
                id_tipo_contrato=documento,
                id_proyecto_consultor=proyectoConsultor
            )
            contrato.save()
                
            persona = proyectoConsultor.id_consultor.id_persona
            # print(persona)
            empresa = proyectoConsultor.id_proyecto.id_empresa_proyecto.id_empresa
            consultor = proyectoConsultor.id_consultor
            usuario = Usuarios.objects.get(id_persona=proyectoConsultor.id_consultor.id_persona)
            dataBancaria = DatosBancarios.objects.get(id_usuario_id=usuario.id)
            contrato = get_object_or_404(Contratos, pk=contrato.id)
            datosbancarios = get_object_or_404(DatosBancarios, pk=dataBancaria.id)
            bancos = get_object_or_404(Bancos, pk=dataBancaria.id_banco.id)

            # Pasar los datos al HTML que servirá como plantilla para el PDF
            template_path = "contratos/contrato_base.html"
            context = {
                "personas":persona,
                "empresas":empresa,
                "consultores":consultor,
                "contratos":contrato,
                "datos_bancarios":datosbancarios,
                "bancos":bancos,
                "fechaInicio":fechaInicio,
                "fechaFinal":fechaFinal,
                "tarifa_diaria":tarifa_diaria.upper(),
                "fechaFirma":fechaFirma,
                "puesto":puesto,
            }

            template = get_template(template_path)
            html = template.render(context)

            # Crear un objeto de respuesta Django y especificar content_type como PDF
            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = 'filename="CONTRATO-BASE.pdf"'


            pdf_bytes = BytesIO()
            # Generar el PDF en memoria usando pisa
            pisa_status = pisa.CreatePDF(html, dest=pdf_bytes, encoding='UTF-8', page_size='letter')

            # Verificar si hubo errores al generar el PDF
            if pisa_status.err:
                raise Exception("Error al generar el PDF")
            
            if consultorCopia == '1':
                endemail(str(usuario.correo), pdf_bytes, consultor.id_persona.nombre, empresa.empresa, proyectoConsultor.id_proyecto.proyecto_nombre, 'CONTRATO BASE', 'GNOSIS [CONTRATO BASE]', 'consultor')

                pass

        
            if empresaCopia == '1':
                sendemail(str(empresa.id_usuario.correo), pdf_bytes, empresa.empresa, consultor.id_persona.nombre, proyectoConsultor.id_proyecto.proyecto_nombre, 'CONTRATO BASE', 'GNOSIS [CONTRATO BASE]', 'empresa')

                pass
            
            # Crear el PDF a partir del HTML usando pisa
            pisa_status = pisa.CreatePDF(html, dest=response, encoding='UTF-8', page_size='letter')

            # Si hay errores al generar el PDF, mostrar un mensaje
            if pisa_status.err:
                return HttpResponse("Error al generar el PDF")

            return response

        except DatosBancarios.DoesNotExist as error:
            print(f"Error: {str(error)}")
            # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
            messages.error(
                request, 'Ha ocurrido un error al procesar la solicitud. Parece que el consultor aun no tiene un registro de sus datos bancarios o de su informacion extra')
            url = reverse('contratoConsultor', args=[consultor.id, proyecto])
            return redirect(url)

        except ProyectoConsultor.DoesNotExist as error:
            print(f"Error: {str(error)}")
            # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
            messages.error(
                request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
            url = reverse('contratoConsultor', args=[consultor.id, proyecto])
            return redirect(url)
    
    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        url = reverse('contratoConsultor', args=[consultor.id, proyecto])
        return redirect(url)
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        url = reverse('contratoConsultor', args=[consultor.id, proyecto])
        return redirect(url)



def sendemail(email, pdf_bytes, nameEmpresa, nameConsultor, proyecto, contrato, titulo, val):
    try:
        to = email
        name = 'Gnosis'
        # Gnosis [REPORTE FINAL DE ACTIVIDADES]"
        context = {
            "nameEmpresa": nameEmpresa,
            "consultorName": nameConsultor,
            "proyectoName": proyecto,
            "contrato": contrato,
            "val":val
        }
        html_content = render_to_string("emails/contratos_email.html", context)

        # Crear el objeto EmailMultiAlternatives para enviar el correo
        email = EmailMultiAlternatives(
            titulo,
            strip_tags(html_content),
            settings.EMAIL_HOST_USER,
            [to],
        )
        fileName = str(contrato) + '.pdf'
        # Adjuntar el PDF al correo
        email.attach(fileName, pdf_bytes.getvalue(), "application/pdf")

        # Indicar que el contenido del correo es HTML
        email.attach_alternative(html_content, "text/html")

        # Enviar el correo
        email.send()

        return "Correo enviado"
    
    except Exception as e:
        print(e)
        return "Correo fallido"



@login_required
def contratoMarco(request):
    try:
        consultor = request.POST.get('consultor_marco')
        proyecto = request.POST.get('proyecto_marco')
        proyectoConsultor = request.POST.get('proyectoConsultor_marco')
        titulo = request.POST.get('titulo_marco')
        fecha_inicio = request.POST.get('fecha_inicio_marco')
        fecha_final = request.POST.get('fecha_final_marco')
        fecha_firma = request.POST.get('fecha_firma_marco')
        atention = request.POST.get('atention')
        person_copy = request.POST.get('person_copy')
        
        empresaCopia = request.POST.get('empresaCopia')
        consultorCopia = request.POST.get('consultorCopia')

        # Asignar un valor por defecto si la variable es None
        empresaCopia = empresaCopia if empresaCopia is not None else '0'
        consultorCopia = consultorCopia if consultorCopia is not None else '0'

        fechaInicio = formatearFecha(fecha_inicio)
        fechaFinal = formatearFecha(fecha_final)
        fechaFirma = formatearFecha(fecha_firma)
        vigencia = tiempoVigencia(request.POST.get('fecha_inicio_marco'), request.POST.get('fecha_final_marco'))
        try:
            proyectoConsultor = ProyectoConsultor.objects.get(pk=int(proyectoConsultor), id_proyecto_id=int(proyecto), id_consultor_id=int(consultor))
            empresa = proyectoConsultor.id_proyecto.id_empresa_proyecto.id_empresa
            titulo = titulo + str(empresa.empresa)
            documento, created = TipoContrato.objects.get_or_create(nombre=titulo)


            try:
            
                contrato = Contratos.objects.get(id_proyecto_consultor=proyectoConsultor, id_tipo_contrato__id=documento.id)

                contrato.fecha_firma=fecha_firma
                contrato.fecha_inicio=fecha_inicio
                contrato.fecha_fin=fecha_final
                contrato.titulo=titulo
                contrato.tiempo_vigencia=vigencia
                contrato.id_tipo_contrato=documento
                contrato.save()

            except Contratos.DoesNotExist:

                contrato = Contratos(
                    titulo=titulo,
                    fecha_firma=fecha_firma,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_final,
                    tiempo_vigencia=vigencia,
                    id_tipo_contrato=documento,
                    id_proyecto_consultor=proyectoConsultor
                )
                contrato.save()

            
            empresa = proyectoConsultor.id_proyecto.id_empresa_proyecto.id_empresa
            usuario = Usuarios.objects.get(id_persona=proyectoConsultor.id_consultor.id_persona)

            consultorEnProyecto = Consultores.objects.get(id_persona=usuario.id_persona)

            dataBancaria = DatosBancarios.objects.get(id_usuario_id=usuario.id)
            dataBancariaEmpresa = DatosBancarios.objects.get(id_usuario_id=empresa.id_usuario)

            contrato = get_object_or_404(Contratos, pk=contrato.id)
            
            datosbancarios = get_object_or_404(DatosBancarios, pk=dataBancaria.id)
            
            bancos = get_object_or_404(Bancos, pk=dataBancaria.id_banco.id)
            
            # Pasar los datos al HTML que servirá como plantilla para el PDF
            template_path = "contratos/marco_gnosis.html"
            context = {
                "empresas":empresa,
                "contratos":contrato,
                "datos_bancarios":datosbancarios,
                "bancos":bancos,
                "dataBancariaEmpresa":dataBancariaEmpresa,
                "atencion":atention,
                "copia":person_copy,
                "fechaFirma":fechaFirma,
                "titulo":titulo
            }

            template = get_template(template_path)
            html = template.render(context)

            # Crear un objeto de respuesta Django y especificar content_type como PDF
            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = f'filename="{titulo}.pdf"'


            pdf_bytes = BytesIO()
            # Generar el PDF en memoria usando pisa
            pisa_status = pisa.CreatePDF(html, dest=pdf_bytes, encoding='UTF-8', page_size='letter')

            # Verificar si hubo errores al generar el PDF
            if pisa_status.err:
                raise Exception("Error al generar el PDF")
            
            if consultorCopia == '1':
                sendemail(str(consultorEnProyecto.correo), pdf_bytes, consultorEnProyecto.id_persona.nombre, empresa.empresa, proyectoConsultor.id_proyecto.proyecto_nombre, 'CONTRATO MARCO GNOSIS', 'GNOSIS [CONTRATO MARCO GNOSIS]', 'consultor')
                
        
            if empresaCopia == '1':
                sendemail(str(empresa.id_usuario.correo), pdf_bytes, empresa.empresa, consultorEnProyecto.id_persona.nombre, proyectoConsultor.id_proyecto.proyecto_nombre, 'CONTRATO MARCO GNOSIS', 'GNOSIS [CONTRATO MARCO GNOSIS]', 'empresa')
                pass

            # Crear el PDF a partir del HTML usando pisa
            pisa_status = pisa.CreatePDF(html, dest=response, encoding='UTF-8', page_size='letter')

            # Si hay errores al generar el PDF, mostrar un mensaje
            if pisa_status.err:
                return HttpResponse("Error al generar el PDF")

            
            return response

        except DatosBancarios.DoesNotExist as error:
            print(f"Error: {str(error)}")
            # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
            messages.error(
                request, 'Ha ocurrido un error al procesar la solicitud. Parece que el consultor o la empresa aun no tienen un registro de sus datos bancarios o de su informacion extra')
            url = reverse('contratoConsultor', args=[consultor, proyecto])
            return redirect(url)

        except ProyectoConsultor.DoesNotExist as error:
            print(f"Error: {str(error)}")
            # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
            messages.error(
                request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
            url = reverse('contratoConsultor', args=[consultor, proyecto])
            return redirect(url)
    
    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        url = reverse('contratoConsultor', args=[consultor, proyecto])
        return redirect(url)
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        url = reverse('contratoConsultor', args=[consultor, proyecto])
        return redirect(url)

@login_required
def contratoAnexo(request):
    try:
        consultor = request.POST.get('consultor_anexo')
        proyecto = request.POST.get('proyecto_anexo')
        proyectoConsultor = request.POST.get('proyectoConsultor_anexo')
        titulo = request.POST.get('titulo_anexo')
        fecha_inicio = request.POST.get('fecha_inicio_anexo')
        fecha_final = request.POST.get('fecha_final_anexo')
        fecha_firma = request.POST.get('fecha_firma_anexo')
        tarifa = request.POST.get('Tarifa_anexo')
        gerente = request.POST.get('gerente_anexo')
        cargo = request.POST.get('cargo_anexo')
        act_1 = request.POST.get('act_1')
        act_2 = request.POST.get('act_2')
        act_3 = request.POST.get('act_3')
        act_4 = request.POST.get('act_4')
        descripcion_1 = request.POST.get('descripcion_1')
        descripcion_2 = request.POST.get('descripcion_2')
        descripcion_3 = request.POST.get('descripcion_3')
        descripcion_4 = request.POST.get('descripcion_4')
        
        empresaCopia = request.POST.get('empresaCopia')
        consultorCopia = request.POST.get('consultorCopia')

        # Asignar un valor por defecto si la variable es None
        empresaCopia = empresaCopia if empresaCopia is not None else '0'
        consultorCopia = consultorCopia if consultorCopia is not None else '0'

        cantidad = "$"+tarifa
        tarifa_diaria = convertir_a_letras(cantidad)

        fechaInicio = formatearFecha(fecha_inicio)
        fechaFinal = formatearFecha(fecha_final)
        fechaFirma = formatearFecha(fecha_firma)
        
        vigencia = tiempoVigencia(request.POST.get('fecha_inicio_anexo'), request.POST.get('fecha_final_anexo'))
        try:
            proyectoConsultor = ProyectoConsultor.objects.get(pk=int(proyectoConsultor), id_proyecto_id=int(proyecto), id_consultor_id=int(consultor))
            empresa = proyectoConsultor.id_proyecto.id_empresa_proyecto.id_empresa
            titulo = titulo + str(empresa.empresa)
            documento, created = TipoContrato.objects.get_or_create(nombre=titulo)
            try:
                
                contrato = Contratos.objects.get(id_proyecto_consultor=proyectoConsultor, id_tipo_contrato__id=documento.id)

                contrato.fecha_firma=fecha_firma
                contrato.fecha_inicio=fecha_inicio
                contrato.fecha_fin=fecha_final
                contrato.titulo=titulo
                contrato.tiempo_vigencia=vigencia
                contrato.id_tipo_contrato=documento
                contrato.tarifa_dia_consultor=tarifa
                contrato.gerente=gerente
                contrato.save()

            except Contratos.DoesNotExist:

                contrato = Contratos(
                    titulo=titulo,
                    fecha_firma=fecha_firma,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_final,
                    tiempo_vigencia=vigencia,
                    tarifa_dia_consultor=tarifa,
                    id_tipo_contrato=documento,
                    gerente=gerente,
                    id_proyecto_consultor=proyectoConsultor
                )
                contrato.save()

            empresa = proyectoConsultor.id_proyecto.id_empresa_proyecto.id_empresa
            usuario = Usuarios.objects.get(id_persona=proyectoConsultor.id_consultor.id_persona)
            dataBancaria = DatosBancarios.objects.get(id_usuario_id=usuario.id)
            consultor = Consultores.objects.get(id_persona=usuario.id_persona)
            contrato = get_object_or_404(Contratos, pk=contrato.id)
            
            datosbancarios = get_object_or_404(DatosBancarios, pk=dataBancaria.id)
            
            bancos = get_object_or_404(Bancos, pk=dataBancaria.id_banco.id)
            
            
            # Pasar los datos al HTML que servirá como plantilla para el PDF
            template_path = "contratos/gnosis_anexo.html"
            context = {
                "fechaFirma":fechaFirma,
                "empresas":empresa,
                "usuarios":usuario,
                "consultor":consultor,
                "fechaInicio":fechaInicio,
                "fechaFinal":fechaFinal,
                "contratos":contrato,
                "cargo":cargo,
                "act_1":act_1,
                "act_2":act_2,
                "act_3":act_3,
                "act_4":act_4,
                "descripcion_1":descripcion_1,
                "descripcion_2":descripcion_2,
                "descripcion_3":descripcion_3,
                "descripcion_4":descripcion_4
            }

            template = get_template(template_path)
            html = template.render(context)

            # Crear un objeto de respuesta Django y especificar content_type como PDF
            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = f'filename="{titulo}.pdf"'


            pdf_bytes = BytesIO()
            # Generar el PDF en memoria usando pisa
            pisa_status = pisa.CreatePDF(html, dest=pdf_bytes, encoding='UTF-8', page_size='letter')

            # Verificar si hubo errores al generar el PDF
            if pisa_status.err:
                raise Exception("Error al generar el PDF")
            

            if consultorCopia == '1':
                sendemail(str(usuario.correo), pdf_bytes, consultor.id_persona.nombre, empresa.empresa, proyectoConsultor.id_proyecto.proyecto_nombre, 'CONTRATO ANEXO', 'GNOSIS [CONTRATO ANEXO]', 'consultor')
                pass
        
            if empresaCopia == '1':
                sendemail(str(empresa.id_usuario.correo), pdf_bytes, empresa.empresa, consultor.id_persona.nombre, proyectoConsultor.id_proyecto.proyecto_nombre, 'CONTRATO ANEXO', 'GNOSIS [CONTRATO ANEXO]', 'empresa')
                pass

            # Crear el PDF a partir del HTML usando pisa
            pisa_status = pisa.CreatePDF(html, dest=response, encoding='UTF-8', page_size='letter')

            # Si hay errores al generar el PDF, mostrar un mensaje
            if pisa_status.err:
                return HttpResponse("Error al generar el PDF")

            return response

        except DatosBancarios.DoesNotExist as error:
            print(f"Error: {str(error)}")
            # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
            messages.error(
                request, 'Ha ocurrido un error al procesar la solicitud. Parece que el consultor aun no tiene un registro de sus datos bancarios o de su informacion extra')
            url = reverse('contratoConsultor', args=[consultor, proyecto])
            return redirect(url)

        except ProyectoConsultor.DoesNotExist as error:
            print(f"Error: {str(error)}")
            # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
            messages.error(
                request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
            url = reverse('contratoConsultor', args=[consultor, proyecto])
            return redirect(url)
    
    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        url = reverse('contratoConsultor', args=[consultor, proyecto])
        return redirect(url)
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        url = reverse('contratoConsultor', args=[consultor, proyecto])
        return redirect(url)


def tiempoVigencia(fecha_inicio_str, fecha_final_str):
    fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
    fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

    # Calcular la diferencia entre las fechas
    diferencia = fecha_final - fecha_inicio

    # Obtener el número de días de vigencia (la propiedad days de la diferencia)
    dias_vigencia = diferencia.days

    # Si deseas obtener el número total de años, meses y días de vigencia, puedes hacer lo siguiente:
    anios_vigencia = dias_vigencia // 365
    meses_vigencia = (dias_vigencia % 365) // 30
    dias_restantes = dias_vigencia % 30

    # Ahora puedes utilizar estos valores para llenar el campo tiempo_vigencia en tu objeto Contratos
    tiempo_vigencia = f'{anios_vigencia} años, {meses_vigencia} meses, {dias_restantes} días'

    return tiempo_vigencia


def formatearFecha(fecha):
    fecha_obj = datetime.strptime(fecha, "%Y-%m-%d")

    # Define un diccionario para mapear los nombres de los meses
    meses = {
        1: "Enero",
        2: "Febrero",
        3: "Marzo",
        4: "Abril",
        5: "Mayo",
        6: "Junio",
        7: "Julio",
        8: "Agosto",
        9: "Septiembre",
        10: "Octubre",
        11: "Noviembre",
        12: "Diciembre"
    }

    # Formatea la fecha en el formato deseado
    fecha_formateada = f"{fecha_obj.day} de {meses[fecha_obj.month]} de {fecha_obj.year}"

    return fecha_formateada


def convertir_a_letras(cantidad):
    # Eliminar el símbolo del dólar y las comas si existen
    cantidad_limpia = cantidad.replace('$', '').replace(',', '')

    # Convertir a número flotante
    numero = float(cantidad_limpia)

    # Convertir el número a letras
    letras = num2words(numero, lang='es')

    return letras



@login_required
def facturasConsultor(request, id, prj):
    try:
        if request.method == 'GET':
            user = Usuarios.objects.get(correo=request.session.get('username'))
            informationPersonalAdministradorUser = Personas.objects.get(pk=user.id_persona_id)
            
            informationConsultorUser = Consultores.objects.get(pk=int(id))
            proyecto = Proyectos.objects.get(pk=int(prj))
            colaborador = ProyectoConsultor.objects.filter(id_proyecto=proyecto,id_consultor=informationConsultorUser)

            if colaborador:
                id_proyecto_consultor_list = colaborador.values_list('id', flat=True)

                # Obtener todos los contratos asociados a los id_proyecto_consultor
                try:
                    contratos = Contratos.objects.get(id_proyecto_consultor__in=id_proyecto_consultor_list)
                except Contratos.DoesNotExist as error:
                    contratos = None

                facturas = Facturas.objects.filter(id_proyecto_consultor__in=id_proyecto_consultor_list)

                regitrosFacturas = []
                for f in facturas:
                    id = f.id
                    periodo = f.periodo
                    fecha = f.id_documentacion.fecha_creacion
                    directory = str(f.id_documentacion.ruta) + '/' + str(f.id_documentacion.nombre)
                    entregado = searchFile(directory)
                    validacionGnosis = f.validacionGnosis
                    validacionEmpresa = f.validacionEmpresa
                    tipoCambio = f.tipoCambioMXN
                    tipoCambioUSD = f.tipoCambioUSD
                    regitrosFacturas.append([periodo, entregado, validacionGnosis, validacionEmpresa, fecha, id, f.id_documentacion.ruta,f.id_documentacion.nombre, tipoCambio,tipoCambioUSD ])


            notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')
        # //Pagination

            # //Cantidad de notificaciones que apareceran antes de crear otra page
            if notification_list.exists():
                paginator = Paginator(notification_list, 6)
                page = request.GET.get('page')
                notification = paginator.get_page(page)

                # // CANT. NOTIFICATIONS PENDING
                pending_total= NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')

            # // ONLY 4 PENDING 
                pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')[:4]
            else:
                notification = []
                pending = []
                pending_total = []

            context = {
                'indice': 'Consultores',
                'contrato':contratos,
                'facturas':regitrosFacturas,
                'notifications':notification,
                'pending':pending,
                'pending_total':pending_total,
            }

            return render(request, 'consultorFacturas.html', context)

        
            

    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Empresas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')


@login_required
def validarFacturas(request):
    try:
        if request.method == 'POST':
            facturaValue = request.POST['factura']
    
            factura = Facturas.objects.get(pk=int(facturaValue))
            
            if 'validar' in request.POST:
                factura.validacionGnosis = 'Aceptada'
                factura.save()
            else:
                factura.validacionGnosis = 'Rechazada'
                factura.save()

            message = 'Ya puedes consultar el estado de tu factura de parte de Gnosis para el proyecto ' + str(factura.id_proyecto_consultor.id_proyecto.proyecto_nombre)

            newNotificacion = NotificationConsultor(name='Gnosis SC', email='gnosis@gmail.com', subject='El estado de tu factura', message=message, status='Pending', id_consultor_destinatary_id=factura.id_proyecto_consultor.id_consultor.id)
                
            newNotificacion.save()

            url = reverse('facturasConsultorForAdmin', args=[factura.id_proyecto_consultor.id_consultor.id, factura.id_proyecto_consultor.id_proyecto.id])

            # Redirigir a la URL construida
            return redirect(url)
            
    except Facturas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')




@login_required
def consultorReportes(request, id, prj):
    try:
        if request.method == 'GET':
            user = Usuarios.objects.get(correo=request.session.get('username'))
            informationPersonalAdministradorUser = Personas.objects.get(pk=user.id_persona_id)
            

            informationConsultorUser = Consultores.objects.get(pk=int(id))
            proyecto = Proyectos.objects.get(pk=int(prj))

            colaborador = ProyectoConsultor.objects.filter(id_proyecto=proyecto,id_consultor=informationConsultorUser)

            if colaborador:
                id_proyecto_consultor_list = colaborador.values_list('id', flat=True)

                # Obtener todos los contratos asociados a los id_proyecto_consultor
                contratos = Contratos.objects.filter(id_proyecto_consultor__in=id_proyecto_consultor_list)

                reportesHoras = ReporteHoras.objects.filter(id_proyecto_consultor__in=id_proyecto_consultor_list)

                regitrosReportesHoras = []
                for f in reportesHoras:
                    id = f.id
                    periodo = f.periodo
                    fecha = f.id_documentacion.fecha_creacion
                    directory = str(f.id_documentacion.ruta) + '/' + str(f.id_documentacion.nombre)
                    entregado = searchFile(directory)
                    validacionGnosis = f.validacionGnosis
                    validacionEmpresa = f.validacionEmpresa

                    regitrosReportesHoras.append([periodo, entregado, validacionGnosis, validacionEmpresa, fecha, id, f.id_documentacion.ruta,f.id_documentacion.nombre ])

                reportesFinales = ReporteFinalActividades.objects.filter(id_proyecto_consultor__in=id_proyecto_consultor_list)

                regitrosReportesFinales = []
                for f in reportesFinales:
                    id = f.id
                    periodo = f.periodo
                    fecha = f.id_documentacion.fecha_creacion
                    directory = str(f.id_documentacion.ruta) + '/' + str(f.id_documentacion.nombre)
                    entregado = searchFile(directory)
                    validacionGnosis = f.validacionGnosis
                    validacionEmpresa = f.validacionEmpresa

                    regitrosReportesFinales.append([periodo, entregado, validacionGnosis, validacionEmpresa, fecha, id, f.id_documentacion.ruta,f.id_documentacion.nombre ])


            notification_list = NotificationAdministrador.objects.filter(id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')
        # //Pagination

            # //Cantidad de notificaciones que apareceran antes de crear otra page
            if notification_list.exists():
                paginator = Paginator(notification_list, 6)
                page = request.GET.get('page')
                notification = paginator.get_page(page)

                # // CANT. NOTIFICATIONS PENDING
                pending_total= NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')

            # // ONLY 4 PENDING 
                pending = NotificationAdministrador.objects.filter(status='Pending',id_persona_destinatary=informationPersonalAdministradorUser.id).order_by('-created_at')[:4]
            else:
                notification = []
                pending = []
                pending_total = []

            context = {
                'indice': 'Consultores',
                'contrato':contratos,
                'reporteHoras':regitrosReportesHoras,
                'reportesFinales':regitrosReportesFinales,
                'notifications':notification,
                'pending':pending,
                'pending_total':pending_total,
            }

            return render(request, 'consultorReportes.html', context)

        
            

    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Empresas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')


@login_required
def validarHoras(request):
    try:
        if request.method == 'POST':
            reporteHora = request.POST['reporteHora']
    
            reporte = ReporteHoras.objects.get(pk=int(reporteHora))
            
            if 'validar' in request.POST:
                reporte.validacionGnosis = 'Aceptada'
                reporte.save()

            else:
                reporte.validacionGnosis = 'Rechazada'
                reporte.save()

            message = 'Ya puedes consultar el estado de tu reporte de horas de parte de Gnosis para el proyecto ' + str(reporte.id_proyecto_consultor.id_proyecto.proyecto_nombre)

            newNotificacion = NotificationConsultor(name='Gnosis SC', email='gnosis@gmail.com', subject='El estado de tu reporte', message=message, status='Pending', id_consultor_destinatary_id=reporte.id_proyecto_consultor.id_consultor.id)
                
            newNotificacion.save()

            url = reverse('consultorReportesForAdmin', args=[reporte.id_proyecto_consultor.id_consultor.id, reporte.id_proyecto_consultor.id_proyecto.id])

            # Redirigir a la URL construida
            return redirect(url)
            
    except Facturas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')




@login_required
def validarReporteFinal(request):
    try:
        if request.method == 'POST':
            reporteFinal = request.POST['reporteFinal']
    
            reporte = ReporteFinalActividades.objects.get(pk=int(reporteFinal))
            
            if 'validar' in request.POST:
                reporte.validacionGnosis = 'Aceptada'
                reporte.save()

            else:
                reporte.validacionGnosis = 'Rechazada'
                reporte.save()

            message = 'Ya puedes consultar el estado de tu reporte final de actividades de parte de Gnosis para el proyecto ' + str(reporte.id_proyecto_consultor.id_proyecto.proyecto_nombre)

            newNotificacion = NotificationConsultor(name='Gnosis SC', email='gnosis@gmail.com', subject='El estado de tu reporte final', message=message, status='Pending', id_consultor_destinatary_id=reporte.id_proyecto_consultor.id_consultor.id)
                
            newNotificacion.save()

            url = reverse('consultorReportesForAdmin', args=[reporte.id_proyecto_consultor.id_consultor.id, reporte.id_proyecto_consultor.id_proyecto.id])

            # Redirigir a la URL construida
            return redirect(url)
            
    except Facturas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')



@login_required
def updateNotaConsultor(request):
    try:
        if request.method == 'POST':
            consultor = request.POST['consultor']
            nota = request.POST['nota']

            queryConsultor = Consultores.objects.get(pk=int(consultor))
            try:
                query = NotasGnosisConsultor.objects.get(id_consultor=queryConsultor)
                query.nota = nota
                query.save()

            except NotasGnosisConsultor.DoesNotExist:
                query = NotasGnosisConsultor(nota=nota, id_consultor=queryConsultor)
                query.save()

            url = reverse('miConsultorProfile', args=[queryConsultor.id_persona.id])

            # Redirigir a la URL construida
            return redirect(url)
            
    except Facturas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')



@login_required
def curriculum_vitaeAdmin(request, id):
    try:
        # descripcion = request.POST.get('descripcion') #perfil descripcion
        
        try:
            informationConsultorUser = Consultores.objects.get(pk=id)
            informationPersonalUser = Personas.objects.get(pk=informationConsultorUser.id_persona.id)
            user = Usuarios.objects.get(id_persona=informationPersonalUser)
    
            #SE PIDEN LOS DATOS A LOS MODELOS
            estudios = Estudios.objects.filter (id_consultor = informationConsultorUser ) #educacion
            curso = CursosConsultor.objects.filter  (id_consultor = informationConsultorUser  ) #certificaciones
            modulos = ConocimientosConsultor.objects.filter( id_consultor = informationConsultorUser  )
            idiomas = IdiomasConsultor.objects.filter( id_consultor = informationConsultorUser)
            proyectoConsultor = ProyectoConsultor.objects.filter(id_consultor = informationConsultorUser  )

            # Pasar los datos al HTML que servirá como plantilla para el PDF
            template_path = "CV/curriculum.html"
            context = {
                "usuarios":user,
                "personas":informationPersonalUser,
                "estudios":estudios,
                "curso":curso,
                "consultores":informationConsultorUser,
                "modulos":modulos,
                "idiomas":idiomas,
                "proyectoConsultor":proyectoConsultor,
            }

            template = get_template(template_path)
            html = template.render(context)
            
            # Crear un objeto de respuesta Django y especificar content_type como PDF
            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = 'filename="CURRICULUM-VITAE.pdf"'
            pdf_bytes = BytesIO()
            
            # Generar el PDF en memoria usando pisa
            pisa_status = pisa.CreatePDF(html, dest=pdf_bytes, encoding='UTF-8', page_size='letter')

            # Verificar si hubo errores al generar el PDF
            if pisa_status.err:
                raise Exception("Error al generar el PDF")


            # Crear el PDF a partir del HTML usando pisa
            pisa_status = pisa.CreatePDF(html, dest=response, encoding='UTF-8', page_size='letter')

            # Si hay errores al generar el PDF, mostrar un mensaje
            if pisa_status.err:
                return HttpResponse("Error al generar el PDF")

            return response

        except IdiomasConsultor.DoesNotExist as error:
            print(f"Error: {str(error)}")
            return redirect('principalAdmin')

        except ProyectoConsultor.DoesNotExist as error:
            print(f"Error: {str(error)}")
            return redirect('principalAdmin')
    
    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        return redirect('principalAdmin')
    except Exception as e:
        print(e)
        return redirect('principalAdmin')




@login_required
def controlDePagosProyectoConsultor(request):
    try:
        tipoCambio()
        cabecera = [
            ['', '', '', '', '', '', '', '', '', 'USD', '', '', '', '', '', '', '', '', '', 'MXP', '', '', ''],
            ['Nombre', 'Módulo', 'Nivel', 'Empresa de', 'Periodo de', 'Fecha', 'Periodo', 'Días', 'Tarifa', 'Cantidad', 'Fecha', 'Fecha', 'Tipo de ', 'Tarifa', 'Tipo de', '', '', '', '', 'Total', 'Días', 'Fecha', 'Mes'],
            ['Consultor', 'SAP', 'SA', 'Asignación', 'Asignación', 'Factura', 'Facturado', 'Facturados', 'Cliente', 'Facturada', 'Promesa', 'Cobranza', 'Cambio', 'Consultor', 'Cambio', 'Subtotal', 'IVA', 'IVA ret', 'ISR ret', 'Pago consultor', 'Pagados', 'Pago', 'Declaración'],
            
        ]

        consultoresProyecto = ProyectoConsultor.objects.all()
        IVA = 0.16
        IVA_Ret = 0.1
        ISR_Ret = 0.1
        
        nombreRegistro = ''

        data = []
        for consultores in consultoresProyecto:
            facturas = Facturas.objects.filter(id_proyecto_consultor=consultores)
            nombre = str(consultores.id_consultor.id_persona.nombre) + ' ' + str(consultores.id_consultor.
            id_persona.ape_pat) + ' ' + str(consultores.id_consultor.id_persona.ape_mat)
            
            
            modulo = str(consultores.id_consultor.especialidad)
            nivel = str(consultores.id_consultor.id_nivel.nombre)
            empresa = str(consultores.id_proyecto.id_empresa_proyecto.id_empresa.empresa)
            periodoAsignacion = formatearFecha(str(consultores.fecha_inicio)) + ' al ' + formatearFecha(str(consultores.fecha_termino))
                    
            # Obtener todas las fechas de factura y periodos asociados al consultor actual
            fechas_factura = [factura.id_documentacion.fecha_creacion for factura in facturas]
            periodos_facturados = [factura.periodo for factura in facturas]
            fechaPago = [factura.fecha_pago for factura in facturas]
            fechaCobranza = [factura.fecha_cobranza for factura in facturas]
            mesDeclarado = [factura.num_mes_declarado for factura in facturas]
            
            try:
                contrato = Contratos.objects.get(id_proyecto_consultor=consultores, titulo='CONTRATO BASE')
                tarifaCliente = contrato.tarifa_dia_consultor
                tarifaCliente = round(float(tarifaCliente), 2)
                fecha_promesa = contrato.fecha_promesa.strftime('%d-%m-%Y')
            except Contratos.DoesNotExist as error:
                tarifaCliente = 0
                fecha_promesa = ''
            
            
            
            tipo_cambio_uno = 10.4645
            tarifa_consultor = int(consultores.tarifa) 
            tipo_cambio_dos = 11.7248
            

            # Verificar si hay la misma cantidad de fechas de factura y periodos antes de combinarlos
            if len(fechas_factura) == len(periodos_facturados):
                for fecha_factura, periodo_facturado, fechaPago, fechaCobranza, mesDeclarado in zip(fechas_factura, periodos_facturados, fechaPago, fechaCobranza, mesDeclarado):
                    diasFacturados = obtener_dias_del_mes(obtener_numero_mes(str(periodo_facturado)), datetime.now().year)
                    dias_pagados = diasFacturados 
                    cantidad_usd_facturada = tarifaCliente*diasFacturados
                    subtotal = dias_pagados*tipo_cambio_dos*tarifa_consultor
                    iva = subtotal*IVA
                    iva = round(iva,3)
                    ivaRet = subtotal*IVA_Ret
                    ivaRet = round(ivaRet,3)
                    isrRet = subtotal*ISR_Ret
                    isrRet = round(isrRet,3)
                    mxpTotalPagoConsultor = subtotal+iva-ivaRet-isrRet
                    mxpTotalPagoConsultor = round(mxpTotalPagoConsultor, 3)        

                    fecha_pago = fechaPago.strftime('%d-%m-%Y')
                    fecha_cobranza = fechaCobranza.strftime('%d-%m-%Y')
                    mes_declaracion = mesDeclarado

                    if nombreRegistro != nombre:
                        nombreRegistro = nombre
                        data.append([nombre, modulo, nivel, empresa, periodoAsignacion, fecha_factura.strftime('%d-%m-%Y'), periodo_facturado, diasFacturados, tarifaCliente, cantidad_usd_facturada, fecha_promesa, fecha_cobranza, tipo_cambio_uno, tarifa_consultor, tipo_cambio_dos, subtotal, iva, ivaRet, isrRet, mxpTotalPagoConsultor, dias_pagados, fecha_pago, mes_declaracion ])

                    else:
                        data.append(['', '', '', '', '', fecha_factura.strftime('%d-%m-%Y'), periodo_facturado, diasFacturados, tarifaCliente, cantidad_usd_facturada, fecha_promesa, fecha_cobranza, tipo_cambio_uno, tarifa_consultor, tipo_cambio_dos, subtotal, iva, ivaRet, isrRet, mxpTotalPagoConsultor, dias_pagados, fecha_pago, mes_declaracion ])
                    
            else:
                # Manejar el caso en el que las listas no tengan la misma longitud (opcional)
                print("Error: No hay coincidencia entre fechas de factura y periodos para el consultor:", nombre)
            
            
        

        return HttpResponse(status=200)

        
    except Facturas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')


def excel(request):
    try:
        cabecera = [
            ['', '', '', '', '', '', '', '', '', 'USD', '', '', '', '', '', '', '', '', '', 'MXP', '', '', ''],
            ['Nombre', 'Módulo', 'Nivel', 'Empresa de', 'Periodo de', 'Fecha', 'Periodo', 'Días', 'Tarifa', 'Cantidad', 'Fecha', 'Fecha', 'Tipo de ', 'Tarifa', 'Tipo de', '', '', '', '', 'Total', 'Días', 'Fecha', 'Mes'],
            ['Consultor', 'SAP', 'SA', 'Asignación', 'Asignación', 'Factura', 'Facturado', 'Facturados', 'Cliente', 'Facturada', 'Promesa', 'Cobranza', 'Cambio', 'Consultor', 'Cambio', 'Subtotal', 'IVA', 'IVA ret', 'ISR ret', 'Pago consultor', 'Pagados', 'Pago', 'Declaración'],
   
        ]

        nombreRegistro = ''
        consultoresProyecto = ProyectoConsultor.objects.all()
        IVA = Decimal('0.16')
        IVA_Ret = Decimal('0.1')
        ISR_Ret = Decimal('0.1')

        data = []
        for consultores in consultoresProyecto:
            facturas = Facturas.objects.filter(id_proyecto_consultor=consultores)
            nombre = str(consultores.id_consultor.id_persona.nombre) + ' ' + str(consultores.id_consultor.
            id_persona.ape_pat) + ' ' + str(consultores.id_consultor.id_persona.ape_mat)
            

            modulo = str(consultores.id_consultor.especialidad)
            nivel = str(consultores.id_consultor.id_nivel.nombre)
            empresa = str(consultores.id_proyecto.id_empresa_proyecto.id_empresa.empresa)
            periodoAsignacion = formatearFecha(str(consultores.fecha_inicio)) + ' al ' + formatearFecha(str(consultores.fecha_termino))
                    
            # Obtener todas las fechas de factura y periodos asociados al consultor actual
            fechas_factura = [factura.id_documentacion.fecha_creacion for factura in facturas]
            periodos_facturados = [factura.periodo for factura in facturas]
            fechaPago = [factura.fecha_pago for factura in facturas]
            fechaCobranza = [factura.fecha_cobranza for factura in facturas]
            mesDeclarado = [factura.num_mes_declarado for factura in facturas]
            tipo_cambio_unoOpcional = [factura.tipoCambioUSD for factura in facturas]
            cambioOpcional = [factura.tipoCambioMXN for factura in facturas]
            

            try:
                titulos_validos = ['CONTRATO BASE', 'CONTRATO FINAL TERCEROS', 'CONTRATO FINAL PROPIO']
                contrato = Contratos.objects.get(id_proyecto_consultor=consultores, titulo__in=titulos_validos)
                tarifaCliente = contrato.tarifa_dia_cliente
                tarifaCliente = round(float(tarifaCliente), 2)
                # print(contrato)
                tarifa_consultor = contrato.tarifa_dia_consultor
                tarifa_consultor = round(float(tarifa_consultor), 2)
                fecha_promesa = contrato.fecha_promesa.strftime('%d-%m-%Y')
                tipo_cambio_uno = contrato.tipoCambioAuto
                
                if tipo_cambio_uno is None:
                    tipo_cambio_uno = 0

                tipo_cambio_dos = contrato.tipoCambioManual
                if tipo_cambio_dos is None:
                    tipo_cambio_dos = 0
                
                
                if contrato.aplicada == 'Automatica':
                    aplicadaCambio = 1
                elif contrato.aplicada == 'Dia - Factura':
                    aplicadaCambio = 2
                    
                else:
                    aplicadaCambio = 0
                cambio = contrato.tipoCambio

                aplicada = contrato.aplicada
            except Contratos.DoesNotExist as error:
                fecha_promesa = ''
            
            
            """if aplicadaCambio == 2:
                tipo_cambio_uno = [factura.tipoCambioUSD for factura in facturas]
                cambio = [factura.tipoCambioMXN for factura in facturas]"""

            # Verificar si hay la misma cantidad de fechas de factura y periodos antes de combinarlos
            if len(fechas_factura) == len(periodos_facturados):
                # Agregar los datos del consultor junto con las fechas de factura y periodos a la lista data
                for fecha_factura, periodo_facturado, fechaPago, fechaCobranza, mesDeclarado, tipo_cambio_unoOpcional,cambioOpcional  in zip(fechas_factura, periodos_facturados, fechaPago, fechaCobranza, mesDeclarado, tipo_cambio_unoOpcional,cambioOpcional ):

                    if aplicadaCambio == 2:
                        cambio = cambioOpcional

                    diasFacturados = obtener_dias_del_mes(obtener_numero_mes(str(periodo_facturado)), datetime.now().year)
                    dias_pagados = diasFacturados 
                    # print("---")
                    # print(tarifaCliente)
                    # print(type(tarifaCliente))
                    # print(diasFacturados)
                    # print(type(diasFacturados))
                    # print(cambio)
                    # print(type(cambio))
                    tarifa_cliente_decimal = Decimal(str(tarifaCliente))

                    # Realizar la multiplicación usando objetos Decimal
                    cantidad_usd_facturada = tarifa_cliente_decimal * diasFacturados * cambio
                    # print(cantidad_usd_facturada)
                    # print("---")
                    # print(dias_pagados)
                    # print(type(tipo_cambio_dos))
                    # print(tarifa_consultor)
                    tarifa_consultor = Decimal(str(tarifa_consultor))

                    if aplicadaCambio == 1:
                        # print("auot")
                        subtotal = dias_pagados*tipo_cambio_uno*tarifa_consultor
                    elif aplicadaCambio == 2:
                        # print("al dia")
                        tipo_cambio_uno = tipo_cambio_unoOpcional
                        subtotal = dias_pagados*tipo_cambio_uno*tarifa_consultor
                    else:
                        # print("manual")
                        subtotal = dias_pagados*tipo_cambio_dos*tarifa_consultor


                    # print("****+")
                    # print(subtotal)
                    iva = subtotal*IVA
                    iva = round(iva,3)
                    ivaRet = subtotal*IVA_Ret
                    ivaRet = round(ivaRet,3)
                    isrRet = subtotal*ISR_Ret
                    isrRet = round(isrRet,3)
                    mxpTotalPagoConsultor = subtotal+iva-ivaRet-isrRet
                    mxpTotalPagoConsultor = round(mxpTotalPagoConsultor, 3)        

                    fecha_pago = fechaPago.strftime('%d-%m-%Y')
                    fecha_cobranza = fechaCobranza.strftime('%d-%m-%Y')
                    mes_declaracion = mesDeclarado

                    if nombreRegistro != nombre:
                        nombreRegistro = nombre
                        data.append([nombre, modulo, nivel, empresa, periodoAsignacion, fecha_factura.strftime('%d-%m-%Y'), periodo_facturado, diasFacturados, tarifaCliente, cantidad_usd_facturada, fecha_promesa, fecha_cobranza, tipo_cambio_uno, tarifa_consultor, tipo_cambio_dos, subtotal, iva, ivaRet, isrRet, mxpTotalPagoConsultor, dias_pagados, fecha_pago, mes_declaracion ])

                    else:
                        data.append(['', '', '', '', '', fecha_factura.strftime('%d-%m-%Y'), periodo_facturado, diasFacturados, tarifaCliente, cantidad_usd_facturada, fecha_promesa, fecha_cobranza, tipo_cambio_uno, tarifa_consultor, tipo_cambio_dos, subtotal, iva, ivaRet, isrRet, mxpTotalPagoConsultor, dias_pagados, fecha_pago, mes_declaracion ])
                    
            else:
                # Manejar el caso en el que las listas no tengan la misma longitud (opcional)
                print("Error: No hay coincidencia entre fechas de factura y periodos para el consultor:", nombre)
                
            
            
        # Crear el libro de trabajo y seleccionar la hoja activa
        wb = Workbook()
        ws = wb.active

        # Definir estilos de formato
        bold_font = Font(bold=True)
        gray_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Agregar las cabeceras a la hoja de cálculo con los estilos aplicados
        for fila in cabecera:
            ws.append(fila)

        # Obtener el número de la última fila con datos en la hoja de trabajo
        last_row = ws.max_row

        # Verificar si la hoja de cálculo contiene datos antes de aplicar los estilos
        if last_row > 0:
            for row in ws.iter_rows(min_row=0, max_row=last_row):
                for cell in row:
                    cell.font = bold_font
                    cell.fill = gray_fill
                    cell.alignment = center_alignment

        # Agregar los datos a la hoja de cálculo
        for fila in data:
            ws.append(fila)

        # Crear una respuesta de Django con el contenido del archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Establecer el encabezado para descargar el archivo con un nombre específico
        response['Content-Disposition'] = 'attachment; filename=datos_ejemplo1.xlsx'
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length
        
        # Crear el objeto de estilo para los bordes
        border_style = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Recorrer todas las celdas y aplicar los bordes
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border_style
        
        ws.title = "Facturación"
        # Guardar el contenido del libro de trabajo en la respuesta HTTP
        wb.save(response)

        # Devolver la respuesta con el archivo Excel
        return response

        
    except Facturas.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')
    
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        return redirect('logout')


def convertir_mxn_a_usd(cantidad_mxn):
    c = CurrencyRates()
    tasa_cambio_mxn_usd = c.get_rate('MXN', 'USD')
    cantidad_usd = cantidad_mxn / tasa_cambio_mxn_usd
    numero_redondeado = round(cantidad_usd, 3)
    return numero_redondeado

    # Convertir $1000 MXN a dólares estadounidenses (USD)

def obtener_dias_del_mes(number_mes, year):
    try:
        num_days = monthrange(year, number_mes)[1] 
        return num_days
    except ValueError:
        return None

def obtener_numero_mes(nombre_mes):
    meses = {
        'ENERO': 1,
        'FEBRERO': 2,
        'MARZO': 3,
        'ABRIL': 4,
        'MAYO': 5,
        'JUNIO': 6,
        'JULIO': 7,
        'AGOSTO': 8,
        'SEPTIEMBRE': 9,
        'OCTUBRE': 10,
        'NOVIEMBRE': 11,
        'DICIEMBRE': 12
    }

    # Convertir el nombre del mes a mayúsculas para evitar errores de coincidencia
    nombre_mes = nombre_mes.upper()

    # Obtener el número del mes del diccionario
    numero_mes = meses.get(nombre_mes)

    return numero_mes


def calcular_iva(monto_base, tasa_iva):
    iva = monto_base * (tasa_iva / 100)
    return iva




@login_required
def contratoFinal(request):
    try:
        consultor = request.POST.get('consultor')
        proyecto = request.POST.get('proyecto')
        proyectoConsultor = request.POST.get('proyectoConsultor')
        titulo = request.POST.get('titulo')
        fecha_inicio = request.POST.get('fecha_inicio_final')
        fecha_final = request.POST.get('fecha_final_final')
        fecha_firma = request.POST.get('fecha_firma_final')
        tarifa = request.POST.get('tarifa_final')
        tarifaCliente = request.POST.get('tarifa_final-cliente')
        gratificacion = request.POST.get('gratificacion')
        moneda = request.POST.get('moneda')
        puesto = request.POST.get('puesto')

        manual = request.POST.get('manual')
        cambioManual = request.POST.get('tipoCambioHoy-Manual')
        cambioAuto = request.POST.get('tipoCambioHoy-Auto')
        tipoCambio = request.POST.get('cambio')

        facturaDia = request.POST.get('factura-dia-propio')
        
        if facturaDia == 'ON':
            # print("factura dia")
            aplicada = 'Dia - Factura'
            cambioAuto = 0.00
            cambioManual = 0.00
            tipoCambio = 0.00
        else:

            if manual == '1':
                # print("manual")
                aplicada = 'Manual'
                
            else:
                # print("api")
                aplicada = 'Automatica'
            
        # print(cambioManual)
        # print(cambioAuto)
        # print(tipoCambio)
        # print("----")

        empresaCopia = request.POST.get('empresaCopia')
        consultorCopia = request.POST.get('consultorCopia')

        # Asignar un valor por defecto si la variable es None
        empresaCopia = empresaCopia if empresaCopia is not None else '0'
        consultorCopia = consultorCopia if consultorCopia is not None else '0'


        monedaValor = TipoMoneda.objects.get(pk=1)
        fechaInicio = formatearFecha(fecha_inicio)
        fechaFinal = formatearFecha(fecha_final)
        fechaFirma = formatearFecha(fecha_firma)
        
        cantidad = "$"+tarifa
        tarifa_diaria = convertir_a_letras(cantidad)
        vigencia = tiempoVigencia(request.POST.get('fecha_inicio_final'), request.POST.get('fecha_final_final'))

        try:
            proyectoConsultor = ProyectoConsultor.objects.get(pk=int(proyectoConsultor), id_proyecto_id=int(proyecto), id_consultor_id=int(consultor))
            # print(proyectoConsultor)
            documento, created = TipoContrato.objects.get_or_create(nombre=titulo)
            
            
            try:
            
                contrato = Contratos.objects.get(id_proyecto_consultor=proyectoConsultor)
                contrato.delete()
                facturas = Facturas.objects.filter(id_proyecto_consultor=proyectoConsultor)
                if facturas:
                    for fac in facturas:
                        fac.delete()

                """
                contrato.fecha_firma=fecha_firma
                contrato.fecha_inicio=fecha_inicio
                contrato.fecha_fin=fecha_final
                contrato.tarifa_dia_consultor=tarifa
                contrato.tarifa_dia_cliente=tarifaCliente
                contrato.tipoCambio=tipoCambio
                contrato.tipoCambioManual=cambioManual
                contrato.tipoCambioAuto=cambioAuto
                contrato.aplicada=aplicada
                contrato.gratificacion=gratificacion
                contrato.tiempo_vigencia=vigencia
                contrato.id_tipo_moneda=monedaValor
                contrato.id_tipo_contrato=documento
                contrato.save()"""

            except Contratos.DoesNotExist:
                pass
                
                
            contrato = Contratos(
                titulo='CONTRATO FINAL PROPIO',
                fecha_firma=fecha_firma,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_final,
                tarifa_dia_consultor=tarifa,
                tarifa_dia_cliente=tarifaCliente,
                gratificacion=gratificacion,
                tiempo_vigencia=vigencia,
                tipoCambio=tipoCambio,
                tipoCambioManual=cambioManual,
                tipoCambioAuto=cambioAuto,
                aplicada=aplicada,
                id_tipo_moneda=monedaValor,
                id_tipo_contrato=documento,
                id_proyecto_consultor=proyectoConsultor
            )
            contrato.save()
                
            persona = proyectoConsultor.id_consultor.id_persona
            # print(persona)
            empresa = proyectoConsultor.id_proyecto.id_empresa_proyecto.id_empresa
            consultor = proyectoConsultor.id_consultor
            usuario = Usuarios.objects.get(id_persona=proyectoConsultor.id_consultor.id_persona)
            dataBancaria = DatosBancarios.objects.get(id_usuario_id=usuario.id)
            contrato = get_object_or_404(Contratos, pk=contrato.id)
            datosbancarios = get_object_or_404(DatosBancarios, pk=dataBancaria.id)
            bancos = get_object_or_404(Bancos, pk=dataBancaria.id_banco.id)
            estudios = Estudios.objects.get(id_consultor=proyectoConsultor.id_consultor)
            proyectos = ProyectoConsultor.objects.get(id_consultor=proyectoConsultor.id_consultor)
            
            if contrato.aplicada == 'Dia - Factura':
                usdPay = contrato.tarifa_dia_consultor
                parte_entera = math.trunc(usdPay)
                parte_decimal = int((usdPay - parte_entera) * 1000)  # Tomando 3 decimales
                numero_truncado = parte_entera + parte_decimal / 1000
                # print(numero_truncado)
                facturaCambio = 1

                usdGratiticado = contrato.gratificacion
                parte_enteraGratificada = math.trunc(usdGratiticado)
                parte_decimalGratificada = int((usdGratiticado - parte_enteraGratificada) * 1000)  # Tomando 3 decimales
                numero_truncadoGratificada = parte_enteraGratificada + parte_decimalGratificada / 1000
                # print(numero_truncadoGratificada)

            else:
                usdPay = contrato.tarifa_dia_consultor*contrato.tipoCambio
                parte_entera = math.trunc(usdPay)
                parte_decimal = int((usdPay - parte_entera) * 1000)  # Tomando 3 decimales
                numero_truncado = parte_entera + parte_decimal / 1000
                # print(numero_truncado)
                facturaCambio = 0

                usdGratiticado = contrato.gratificacion*contrato.tipoCambio
                parte_enteraGratificada = math.trunc(usdGratiticado)
                parte_decimalGratificada = int((usdGratiticado - parte_enteraGratificada) * 1000)  # Tomando 3 decimales
                numero_truncadoGratificada = parte_enteraGratificada + parte_decimalGratificada / 1000
                # print(numero_truncadoGratificada)


            

            # Pasar los datos al HTML que servirá como plantilla para el PDF
            template_path = "contratos/contrato_final_propio.html"
            context = {
                "estudios":estudios,
                "personas":persona,
                "empresas":empresa,
                "consultores":consultor,
                "contratos":contrato,
                "datos_bancarios":datosbancarios,
                "bancos":bancos,
                "fechaInicio":fechaInicio,
                "fechaFinal":fechaFinal,
                "tarifa_diaria":tarifa_diaria.upper(),
                "facturaCambio":facturaCambio,
                "tarifaDiariaUSD": numero_truncado,
                "gratificacion":numero_truncadoGratificada,
                "fechaFirma":fechaFirma,
                "puesto":puesto,
                "proyectos":proyectos,
            }

            template = get_template(template_path)
            html = template.render(context)

            # Crear un objeto de respuesta Django y especificar content_type como PDF
            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = 'filename="CONTRATO-BASE.pdf"'


            pdf_bytes = BytesIO()
            # Generar el PDF en memoria usando pisa
            pisa_status = pisa.CreatePDF(html, dest=pdf_bytes, encoding='UTF-8', page_size='letter')

            # Verificar si hubo errores al generar el PDF
            if pisa_status.err:
                raise Exception("Error al generar el PDF")
            
            if consultorCopia == '1':
                endemail(str(usuario.correo), pdf_bytes, consultor.id_persona.nombre, empresa.empresa, proyectoConsultor.id_proyecto.proyecto_nombre, 'CONTRATO BASE', 'GNOSIS [CONTRATO BASE]', 'consultor')

                pass

        
            if empresaCopia == '1':
                sendemail(str(empresa.id_usuario.correo), pdf_bytes, empresa.empresa, consultor.id_persona.nombre, proyectoConsultor.id_proyecto.proyecto_nombre, 'CONTRATO BASE', 'GNOSIS [CONTRATO BASE]', 'empresa')

                pass
            
            # Crear el PDF a partir del HTML usando pisa
            pisa_status = pisa.CreatePDF(html, dest=response, encoding='UTF-8', page_size='letter')

            # Si hay errores al generar el PDF, mostrar un mensaje
            if pisa_status.err:
                return HttpResponse("Error al generar el PDF")

            return response

        except DatosBancarios.DoesNotExist as error:
            print(f"Error: {str(error)}")
            # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
            messages.error(
                request, 'Ha ocurrido un error al procesar la solicitud. Parece que el consultor aun no tiene un registro de sus datos bancarios o de su informacion extra')
            url = reverse('contratoConsultor', args=[consultor.id, proyecto])
            return redirect(url)

        except ProyectoConsultor.DoesNotExist as error:
            print(f"Error: {str(error)}")
            # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
            messages.error(
                request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
            url = reverse('contratoConsultor', args=[consultor.id, proyecto])
            return redirect(url)
    
    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        url = reverse('contratoConsultor', args=[consultor.id, proyecto])
        return redirect(url)
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        url = reverse('contratoConsultor', args=[consultor.id, proyecto])
        return redirect(url)


@login_required
def contratoTerceros(request):
    try:
        consultor = request.POST.get('consultor')
        proyecto = request.POST.get('proyecto')
        proyectoConsultor = request.POST.get('proyectoConsultor')
        titulo = request.POST.get('titulo')
        fecha_inicio = request.POST.get('fecha_inicio_terceros')
        fecha_final = request.POST.get('fecha_final_terceros')
        fecha_firma = request.POST.get('fecha_firma_terceros')
        tarifa = request.POST.get('tarifa_final_terceros')
        tarifaCliente = request.POST.get('tarifa_final_terceros-cliente')
        moneda = request.POST.get('moneda')
        persona_tercero = request.POST.get('persona_tercero')
        gratificacion = request.POST.get('gratificacion_tercero')
        tipo_recibo = request.POST.get('tipo_recibo')

        empresaCopia = request.POST.get('empresaCopia')
        consultorCopia = request.POST.get('consultorCopia')

        # Asignar un valor por defecto si la variable es None
        empresaCopia = empresaCopia if empresaCopia is not None else '0'
        consultorCopia = consultorCopia if consultorCopia is not None else '0'

        manual = request.POST.get('manual-terceros')
        cambioManual = request.POST.get('tipoCambioHoy-Manual-terceros')
        cambioAuto = request.POST.get('tipoCambioHoy-Auto-terceros')
        tipoCambio = request.POST.get('cambio-terceros')

        facturaDia = request.POST.get('factura-dia')
        
        if facturaDia == 'ON':
            # print("factura dia")
            aplicada = 'Dia - Factura'
            cambioAuto = 0.00
            cambioManual = 0.00
            tipoCambio = 0.00
        else:
            if manual == '1':
                # print("manual")
                aplicada = 'Manual'
                
            else:
                # print("api")
                aplicada = 'Automatica'
                


        monedaValor = TipoMoneda.objects.get(pk=1)
        fechaInicio = formatearFecha(fecha_inicio)
        fechaFinal = formatearFecha(fecha_final)
        fechaFirma = formatearFecha(fecha_firma)
        
        cantidad = "$"+tarifa
        tarifa_diaria = convertir_a_letras(cantidad)
        vigencia = tiempoVigencia(request.POST.get('fecha_inicio_terceros'), request.POST.get('fecha_final_terceros'))

        try:
            proyectoConsultor = ProyectoConsultor.objects.get(pk=int(proyectoConsultor), id_proyecto_id=int(proyecto), id_consultor_id=int(consultor))

            documento, created = TipoContrato.objects.get_or_create(nombre=titulo)
            
            contratacion, val = TipoContrataciones.objects.get_or_create(contratacion=tipo_recibo)
            
            try:
            
                contrato = Contratos.objects.get(id_proyecto_consultor=proyectoConsultor)
                contrato.delete()
                facturas = Facturas.objects.filter(id_proyecto_consultor=proyectoConsultor)
                if facturas:
                    for fac in facturas:
                        fac.delete()
                """
                contrato.fecha_firma=fecha_firma
                contrato.fecha_inicio=fecha_inicio
                contrato.fecha_fin=fecha_final
                contrato.tarifa_dia_consultor=tarifa
                contrato.tarifa_dia_cliente=tarifaCliente
                contrato.tiempo_vigencia=vigencia
                contrato.persona_tercero=persona_tercero
                contrato.gratificacion=gratificacion
                contrato.id_tipo_moneda=monedaValor
                contrato.tipoCambio=tipoCambio
                contrato.tipoCambioManual=cambioManual
                contrato.tipoCambioAuto=cambioAuto
                contrato.aplicada=aplicada
                contrato.id_tipo_contrato=documento
                contrato.id_tipo_contratacion=contratacion
                contrato.save()"""

            except Contratos.DoesNotExist:
                pass

            contrato = Contratos(
                titulo='CONTRATO FINAL TERCEROS',
                fecha_firma=fecha_firma,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_final,
                tarifa_dia_consultor=tarifa,
                tarifa_dia_cliente=tarifaCliente,
                tiempo_vigencia=vigencia,
                tipoCambio=tipoCambio,
                tipoCambioManual=cambioManual,
                tipoCambioAuto=cambioAuto,
                aplicada=aplicada,
                persona_tercero=persona_tercero,
                gratificacion=gratificacion,
                id_tipo_moneda=monedaValor,
                id_tipo_contratacion=contratacion,
                id_tipo_contrato=documento,
                id_proyecto_consultor=proyectoConsultor
            )
            contrato.save()
                
            persona = proyectoConsultor.id_consultor.id_persona
            empresa = proyectoConsultor.id_proyecto.id_empresa_proyecto.id_empresa
            consultor = proyectoConsultor.id_consultor
            usuario = Usuarios.objects.get(id_persona=proyectoConsultor.id_consultor.id_persona)
            dataBancaria = DatosBancarios.objects.get(id_usuario_id=usuario.id)
            contrato = get_object_or_404(Contratos, pk=contrato.id)
            datosbancarios = get_object_or_404(DatosBancarios, pk=dataBancaria.id)
            bancos = get_object_or_404(Bancos, pk=dataBancaria.id_banco.id)
            estudios = Estudios.objects.get(id_consultor=proyectoConsultor.id_consultor)
            proyectos = ProyectoConsultor.objects.get(id_consultor=proyectoConsultor.id_consultor)

            if contrato.aplicada == 'Dia - Factura':
                usdPay = contrato.tarifa_dia_consultor
                # print(contrato.tarifa_dia_consultor)
                # print(contrato.tipoCambio)
                # print(usdPay)
                parte_entera = math.trunc(usdPay)
                parte_decimal = int((usdPay - parte_entera) * 1000)  # Tomando 3 decimales
                numero_truncado = parte_entera + parte_decimal / 1000
                # print(numero_truncado)

                usdGratiticado = contrato.gratificacion
                parte_enteraGratificada = math.trunc(usdGratiticado)
                parte_decimalGratificada = int((usdGratiticado - parte_enteraGratificada) * 1000)  # Tomando 3 decimales
                numero_truncadoGratificada = parte_enteraGratificada + parte_decimalGratificada / 1000
                # print(numero_truncadoGratificada)
                facturaCambio = 1

            else:
                usdPay = contrato.tarifa_dia_consultor*contrato.tipoCambio
                # print(contrato.tarifa_dia_consultor)
                # print(contrato.tipoCambio)
                # print(usdPay)
                parte_entera = math.trunc(usdPay)
                parte_decimal = int((usdPay - parte_entera) * 1000)  # Tomando 3 decimales
                numero_truncado = parte_entera + parte_decimal / 1000
                # print(numero_truncado)

                usdGratiticado = contrato.gratificacion*contrato.tipoCambio
                parte_enteraGratificada = math.trunc(usdGratiticado)
                parte_decimalGratificada = int((usdGratiticado - parte_enteraGratificada) * 1000)  # Tomando 3 decimales
                numero_truncadoGratificada = parte_enteraGratificada + parte_decimalGratificada / 1000
                # print(numero_truncadoGratificada)
                facturaCambio = 0

            # Pasar los datos al HTML que servirá como plantilla para el PDF
            template_path = "contratos/contrato_terceros.html"
            context = {
                "estudios":estudios,
                "personas":persona,
                "empresas":empresa,
                "consultores":consultor,
                "contratos":contrato,
                "datos_bancarios":datosbancarios,
                "bancos":bancos,
                "gratificacion":gratificacion,
                "fechaInicio":fechaInicio,
                "fechaFinal":fechaFinal,
                "tarifaDiariaUSD": numero_truncado,
                "gratificacion":numero_truncadoGratificada,
                "tarifa_diaria":tarifa_diaria.upper(),
                "facturaCambio":facturaCambio,
                "fechaFirma":fechaFirma,
                "proyectos":proyectos,
            }

            template = get_template(template_path)
            html = template.render(context)

            # Crear un objeto de respuesta Django y especificar content_type como PDF
            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = 'filename="CONTRATO-BASE.pdf"'


            pdf_bytes = BytesIO()
            # Generar el PDF en memoria usando pisa
            pisa_status = pisa.CreatePDF(html, dest=pdf_bytes, encoding='UTF-8', page_size='letter')

            # Verificar si hubo errores al generar el PDF
            if pisa_status.err:
                raise Exception("Error al generar el PDF")
            
            if consultorCopia == '1':
                sendemail(str(usuario.correo), pdf_bytes, consultor.id_persona.nombre, empresa.empresa, proyectoConsultor.id_proyecto.proyecto_nombre, 'CONTRATO BASE', 'GNOSIS [CONTRATO BASE]', 'consultor')

                pass

        
            if empresaCopia == '1':
                sendemail(str(empresa.id_usuario.correo), pdf_bytes, empresa.empresa, consultor.id_persona.nombre, proyectoConsultor.id_proyecto.proyecto_nombre, 'CONTRATO BASE', 'GNOSIS [CONTRATO BASE]', 'empresa')

                pass
            
            # Crear el PDF a partir del HTML usando pisa
            pisa_status = pisa.CreatePDF(html, dest=response, encoding='UTF-8', page_size='letter')

            # Si hay errores al generar el PDF, mostrar un mensaje
            if pisa_status.err:
                return HttpResponse("Error al generar el PDF")

            return response

        except DatosBancarios.DoesNotExist as error:
            print(f"Error: {str(error)}")
            # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
            messages.error(
                request, 'Ha ocurrido un error al procesar la solicitud. Parece que el consultor aun no tiene un registro de sus datos bancarios o de su informacion extra')
            url = reverse('contratoConsultor', args=[consultor.id, proyecto])
            return redirect(url)

        except ProyectoConsultor.DoesNotExist as error:
            print(f"Error: {str(error)}")
            # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
            messages.error(
                request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
            url = reverse('contratoConsultor', args=[consultor.id, proyecto])
            return redirect(url)
    
    except Usuarios.DoesNotExist as error:
        print(f"Error: {str(error)}")
        # Error generado por que el usuario no realizo su registro completo, para solucionar el error se debera borrar todo el rastro del usuario, y buscar las relaciones perdidas
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        url = reverse('contratoConsultor', args=[consultor.id, proyecto])
        return redirect(url)
    except Exception as e:
        print(e)
        messages.error(
            request, 'Ha ocurrido un error al procesar la solicitud. Ponte en contacto con soporte tecnico')
        url = reverse('contratoConsultor', args=[consultor.id, proyecto])
        return redirect(url)


def tipoCambioUSD():
    api_key = "93c3b7f4d1f22af42757e75f"
    url = f"https://v6.exchangerate-api.com/v6/{api_key}/latest/MXN"
    response = requests.get(url)

    if response.status_code == 200:
        data = response.json()
        exchange_rates = data["conversion_rates"]
        for currency, rate in exchange_rates.items():
            if currency == 'USD':
                # print(f"1 USD = {round(float(rate),2)} {currency}")
                st = str(rate)
                return st[0:6]
    else:
        print("Error:", response.status_code)
        return None


def tipoCambioMXN():
    api_key = "93c3b7f4d1f22af42757e75f"
    url = f"https://v6.exchangerate-api.com/v6/{api_key}/latest/USD"
    response = requests.get(url)

    if response.status_code == 200:
        data = response.json()
        exchange_rates = data["conversion_rates"]
        for currency, rate in exchange_rates.items():
            if currency == 'MXN':
                # print(f"1 USD = {round(float(rate),2)} {currency}")
                st = str(rate)
                return st[0:6]
    else:
        print("Error:", response.status_code)
        return None